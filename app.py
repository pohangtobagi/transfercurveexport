
import io
import json
import re
import uuid
import pickle
import base64
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd
import streamlit as st

try:
    from supabase import create_client
except ImportError:
    create_client = None
import plotly.graph_objects as go
from plotly.subplots import make_subplots

st.set_page_config(page_title="FET-Analysis_Minjae", layout="wide")
st.title("FET-Analysis_Minjae")

st.markdown("""
<style>
div[data-testid="stSlider"] div[role="slider"] {
    background-color: black !important;
    border-color: black !important;
}
div[data-testid="stSlider"] div[data-testid="stSliderTrack"] > div:nth-child(1) {
    background-color: black !important;
}
div[data-testid="stSlider"] div[role="slider"] > div {
    color: var(--text-color) !important;
    background-color: transparent !important;
}

/* Plot-aligned parameter regions: compact enough to avoid page scrolling */
div[data-testid="stPlotlyChart"] {
    margin-bottom: -58px !important;
}
.parameter-region {
    margin-top: -22px !important;
    min-height: 430px !important;
    overflow: visible !important;
}
.parameter-region div[data-testid="stVerticalBlockBorderWrapper"] {
    border: none !important;
    background: transparent !important;
    box-shadow: none !important;
    padding: 0.18rem 0.12rem !important;
}
.parameter-emphasis {
    font-size: 19px;
    font-weight: 900;
    text-align: left !important;
    padding: 0 2px 5px 2px;
    margin: 0 !important;
    line-height: 1.25 !important;
}
.metric-title {
    font-size: 14px;
    font-weight: 800;
    line-height: 1.18;
    margin-bottom: 6px;
}
.metric-value {
    font-size: 18px;
    font-weight: 800;
    line-height: 1.15;
    white-space: nowrap;
}
div[data-testid="stMainBlockContainer"] div[data-testid="stVerticalBlock"] {
    gap: 0.28rem !important;
}
div[data-testid="stMainBlockContainer"] div[data-testid="stHorizontalBlock"] {
    gap: 0.35rem !important;
}

/* Direction-aware compact summary */
.data-header-row {
    display:flex;
    align-items:center;
    gap:12px;
    margin:2px 0 4px 0;
}
.top-param-card {
    min-height:78px;
    padding:8px 10px;
    border-radius:8px;
    background:rgba(120,120,120,0.055);
}
.top-param-title {
    font-size:13px;
    font-weight:800;
    color:#555;
    margin-bottom:5px;
}
.top-param-value {
    font-size:20px;
    font-weight:850;
    line-height:1.2;
    white-space:nowrap;
    overflow:hidden;
    text-overflow:ellipsis;
}
.compact-slider-area {
    margin-top:-34px;
}
.compact-slider-area div[data-testid="stSelectSlider"] label,
.compact-slider-area div[data-testid="stButton"] button p {
    font-size:10px !important;
}
.direction-caption {
    font-size:12px;
    font-weight:800;
    margin-bottom:2px;
}

/* v42 final layout */
.top-param-card {
    min-height: 72px !important;
    padding: 6px 4px !important;
    border-radius: 0 !important;
    background: transparent !important;
    border: none !important;
}
.top-param-title {
    font-size: 16px !important;
    font-weight: 850 !important;
    color: #444 !important;
    margin-bottom: 4px !important;
}
.top-param-value {
    font-size: 23px !important;
    font-weight: 900 !important;
    line-height: 1.22 !important;
    overflow: visible !important;
    text-overflow: clip !important;
}
.slider-heading {
    font-size: 15px;
    font-weight: 900;
    margin: 0 0 4px 0;
}
.compact-slider-area div[data-testid="stButton"] button {
    min-height: 28px !important;
}

/* v43 layout refinements */
.top-param-title {
    font-size: 19px !important;
    font-weight: 900 !important;
    margin-bottom: 6px !important;
}
.top-param-value {
    font-size: 27px !important;
    font-weight: 900 !important;
}
.param-row-gap {
    height: 18px;
}
.slider-row-gap {
    height: 18px;
}
.slider-heading {
    font-size: 17px !important;
    margin-bottom: 7px !important;
}

/* v44 slider spacing */
.compact-slider-area {
    margin-top: -8px !important;
}
.compact-slider-area div[data-testid="stSelectSlider"] {
    margin-top: 8px !important;
    margin-bottom: 12px !important;
}
.compact-slider-area div[data-testid="stHorizontalBlock"] {
    gap: 0.65rem !important;
}
.compact-slider-area div[data-testid="stButton"] {
    margin-top: 7px !important;
    margin-bottom: 10px !important;
}
.slider-heading {
    padding-top: 8px !important;
    margin-bottom: 10px !important;
}

/* v45 stable slider layout */
.compact-slider-area div[data-testid="stSelectSlider"] {
    margin-top: 12px !important;
    margin-bottom: 18px !important;
}
.compact-slider-area .slider-heading {
    margin-bottom: 14px !important;
}
.compact-slider-area div[data-testid="stButton"] {
    margin-top: 10px !important;
    margin-bottom: 12px !important;
}

/* v47 unified plot-control geometry */
.compact-slider-area div[data-testid="stSelectSlider"] {
    min-height: 42px !important;
    margin-top: 8px !important;
    margin-bottom: 8px !important;
}
.compact-slider-area div[data-testid="stNumberInput"] {
    margin-top: 4px !important;
    margin-bottom: 8px !important;
}
.compact-slider-area div[data-testid="stNumberInput"] input {
    height: 30px !important;
    font-size: 11px !important;
    padding: 2px 6px !important;
}
.compact-slider-area div[data-testid="stButton"] button {
    height: 32px !important;
    min-height: 32px !important;
}
.control-section-spacer { height: 18px; }
.control-placeholder { min-height: 420px; }

/* v48 aligned plot controls */
.compact-slider-area + div[data-testid="stVerticalBlock"] {
    width: 100% !important;
}
.control-section-spacer {
    height: 26px !important;
}
.control-placeholder {
    min-height: 390px !important;
}
.compact-slider-area div[data-testid="stSelectSlider"] {
    min-height: 58px !important;
    margin-top: 10px !important;
    margin-bottom: 10px !important;
}
.compact-slider-area div[data-testid="stNumberInput"] {
    margin-top: 6px !important;
    margin-bottom: 10px !important;
}
.compact-slider-area div[data-testid="stNumberInput"] input {
    height: 40px !important;
    min-height: 40px !important;
}
.compact-slider-area div[data-testid="stButton"] button {
    min-height: 40px !important;
    height: 40px !important;
}
.compact-slider-area .slider-heading {
    min-height: 25px !important;
    margin-bottom: 8px !important;
}

/* v49 SS range and aligned controls */
div[data-testid="stNumberInput"] label p {
    font-size: 11px !important;
    font-weight: 700 !important;
}
div[data-testid="stNumberInput"] input {
    min-height: 36px !important;
    height: 36px !important;
}
.control-section-spacer {
    height: 20px !important;
}
.slider-heading {
    margin-bottom: 10px !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
/* Project radio labels */
section[data-testid="stSidebar"] div[role="radiogroup"] label p {
    font-size: 17px !important;
    font-weight: 700 !important;
}
section[data-testid="stSidebar"] div[role="radiogroup"] label {
    padding-top: 0.35rem !important;
    padding-bottom: 0.35rem !important;
}

/* Compact sidebar controls and saved-log rows */
section[data-testid="stSidebar"] div[data-testid="stButton"] {
    margin-top: -0.18rem !important;
    margin-bottom: -0.22rem !important;
}
section[data-testid="stSidebar"] div[data-testid="stButton"] button {
    min-height: 28px !important;
    height: 28px !important;
    padding: 0 0.42rem !important;
    border-radius: 6px !important;
}
section[data-testid="stSidebar"] div[data-testid="stButton"] button p {
    font-size: 11px !important;
    line-height: 1 !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
section[data-testid="stSidebar"] div[data-testid="stHorizontalBlock"] {
    gap: 0.22rem !important;
}
section[data-testid="stSidebar"] div[data-testid="stCaptionContainer"] {
    margin-top: -0.15rem !important;
    margin-bottom: -0.1rem !important;
}
section[data-testid="stSidebar"] hr {
    margin-top: 0.45rem !important;
    margin-bottom: 0.45rem !important;
}

/* Compact top parameter and inline control rows */
div[data-testid="stMainBlockContainer"] div[data-testid="stButton"] button {
    min-height: 30px;
}

/* Compact Peak Elimination buttons */
.peak-elimination-box div[data-testid="stButton"] button {
    min-height: 24px !important;
    height: 24px !important;
    padding: 0 0.20rem !important;
}
.peak-elimination-box div[data-testid="stButton"] button p {
    font-size: 9px !important;
    line-height: 1 !important;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# Session log / folder helpers
# ============================================================
STORAGE_DIR = Path(__file__).resolve().parent / ".fet_storage"
STORAGE_FILE = STORAGE_DIR / "projects.pkl"


def _default_persistent_state():
    return {
        "folders": {},
        "active_folder": None,
        "active_log_id": None,
        "project_device_settings": {},
        "project_workspaces": {},
    }


@st.cache_resource(show_spinner=False)
def _build_supabase_client(url, key):
    if create_client is None:
        return None
    return create_client(url, key)


def get_supabase_client():
    """Return one cached Supabase client for the whole app process."""
    if create_client is None:
        return None

    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_SERVICE_ROLE_KEY"]
    except Exception:
        return None

    return _build_supabase_client(url, key)


def serialize_persistent_state(data):
    raw = pickle.dumps(data, protocol=pickle.HIGHEST_PROTOCOL)
    return base64.b64encode(raw).decode("ascii")


def deserialize_persistent_state(payload):
    raw = base64.b64decode(payload.encode("ascii"))
    data = pickle.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("Invalid persistent state")
    return data


def normalize_persistent_state(data):
    if not isinstance(data, dict):
        data = _default_persistent_state()
    data.setdefault("folders", {})
    data.setdefault("active_folder", None)
    data.setdefault("active_log_id", None)
    data.setdefault("project_device_settings", {})
    data.setdefault("project_workspaces", {})
    return data


@st.cache_data(show_spinner=False, max_entries=16)
def cached_deserialize_persistent_state(payload):
    return normalize_persistent_state(
        deserialize_persistent_state(payload)
    )


@st.cache_data(show_spinner=False, max_entries=12)
def cached_excel_sheet_names(file_bytes):
    with io.BytesIO(file_bytes) as buffer:
        return pd.ExcelFile(buffer).sheet_names


@st.cache_data(show_spinner=False, max_entries=128)
def cached_read_excel_sheet(file_bytes, sheet_name, header=0):
    with io.BytesIO(file_bytes) as buffer:
        return pd.read_excel(
            buffer,
            sheet_name=sheet_name,
            header=header,
        )


@st.cache_data(show_spinner=False, max_entries=64)
def cached_settings_drain_v(file_bytes, selected_sheet):
    """Cached Settings lookup; avoids reparsing Settings on every rerun."""
    try:
        sheet_names = cached_excel_sheet_names(file_bytes)
        settings_sheet = next(
            (
                name for name in sheet_names
                if name.strip().lower() == "settings"
            ),
            None,
        )
        if settings_sheet is None:
            return None

        raw = cached_read_excel_sheet(
            file_bytes, settings_sheet, header=None
        )
    except Exception:
        return None

    rows, cols = raw.shape
    selected_norm = normalize_excel_label(selected_sheet or "")
    anchors = []

    if selected_norm:
        for row_idx in range(rows):
            for col_idx in range(cols):
                if (
                    normalize_excel_label(
                        raw.iat[row_idx, col_idx]
                    )
                    == selected_norm
                ):
                    anchors.append((row_idx, col_idx))

    if not anchors:
        anchors = [(0, 0)]

    for anchor_row, _ in anchors:
        row_start = max(0, anchor_row - 3)
        row_end = min(rows, anchor_row + 80)
        name_row = None
        start_level_row = None

        for row_idx in range(row_start, row_end):
            token = normalize_excel_label(raw.iat[row_idx, 0])
            if token == "name":
                name_row = row_idx
            elif token in {"startlevel", "start", "level"}:
                start_level_row = row_idx

        if name_row is None or start_level_row is None:
            continue

        drain_v_col = None
        for col_idx in range(cols):
            token = normalize_excel_label(
                raw.iat[name_row, col_idx]
            )
            if token in {
                "drainv", "drainvoltage", "vd", "vds"
            }:
                drain_v_col = col_idx
                break

        if drain_v_col is None:
            continue

        value = pd.to_numeric(
            pd.Series(
                [raw.iat[start_level_row, drain_v_col]]
            ),
            errors="coerce",
        ).iloc[0]
        if pd.notna(value):
            return float(value)

    return None


def load_projects_state():
    """Load permanent state from Supabase; fall back to local file if needed."""
    client = get_supabase_client()
    if client is not None:
        try:
            response = (
                client.table("fet_app_state")
                .select("payload")
                .eq("state_key", "main")
                .limit(1)
                .execute()
            )
            rows = response.data or []
            if rows and rows[0].get("payload"):
                return cached_deserialize_persistent_state(
                    rows[0]["payload"]
                )
        except Exception as exc:
            st.session_state["persistence_warning"] = (
                f"Supabase 불러오기 실패: {exc}"
            )

    # Local fallback is useful for local development, but Streamlit Cloud
    # should use Supabase because its local disk is not permanent.
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    if STORAGE_FILE.exists():
        try:
            with STORAGE_FILE.open("rb") as handle:
                return normalize_persistent_state(pickle.load(handle))
        except Exception:
            pass
    return _default_persistent_state()


def save_projects_state():
    """Persist projects, logs, settings and uploaded Excel bytes to Supabase."""
    payload = {
        "folders": st.session_state.get("analysis_log_folders", {}),
        "active_folder": st.session_state.get("active_log_folder"),
        "active_log_id": st.session_state.get("persistent_active_log_id"),
        "project_device_settings": st.session_state.get(
            "project_device_settings", {}
        ),
        "project_workspaces": st.session_state.get(
            "project_workspaces", {}
        ),
    }

    encoded = serialize_persistent_state(payload)

    # Skip identical remote writes. Slider reruns can otherwise issue many
    # expensive Supabase upserts with exactly the same payload.
    payload_hash = str(hash(encoded))
    if (
        st.session_state.get("_last_persisted_payload_hash")
        == payload_hash
    ):
        return

    client = get_supabase_client()
    if client is not None:
        try:
            (
                client.table("fet_app_state")
                .upsert(
                    {
                        "state_key": "main",
                        "payload": encoded,
                    },
                    on_conflict="state_key",
                )
                .execute()
            )
            st.session_state[
                "_last_persisted_payload_hash"
            ] = payload_hash
            st.session_state["persistence_status"] = (
                "Supabase 저장 완료"
            )
            return
        except Exception as exc:
            st.session_state["persistence_warning"] = (
                f"Supabase 저장 실패: {exc}"
            )

    # Local fallback for development only.
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    temp_file = STORAGE_FILE.with_suffix(".tmp")
    try:
        with temp_file.open("wb") as handle:
            pickle.dump(payload, handle, protocol=pickle.HIGHEST_PROTOCOL)
        temp_file.replace(STORAGE_FILE)
    except Exception:
        if temp_file.exists():
            try:
                temp_file.unlink()
            except Exception:
                pass

def initialize_log_state():
    if "analysis_log_folders" not in st.session_state:
        saved = load_projects_state()
        st.session_state["analysis_log_folders"] = saved["folders"]
        st.session_state["active_log_folder"] = saved["active_folder"]
        st.session_state["persistent_active_log_id"] = saved["active_log_id"]
        st.session_state["project_device_settings"] = saved.get(
            "project_device_settings", {}
        )
        st.session_state["project_workspaces"] = saved.get(
            "project_workspaces", {}
        )

    if "active_log_folder" not in st.session_state:
        st.session_state["active_log_folder"] = None
    if "persistent_active_log_id" not in st.session_state:
        st.session_state["persistent_active_log_id"] = None
    if "project_device_settings" not in st.session_state:
        st.session_state["project_device_settings"] = {}
    if "project_workspaces" not in st.session_state:
        st.session_state["project_workspaces"] = {}
    if "file_uploader_generation" not in st.session_state:
        st.session_state["file_uploader_generation"] = 0
    if "active_file_source" not in st.session_state:
        st.session_state["active_file_source"] = (
            "log" if st.session_state.get("active_file_bytes") else "upload"
        )


def create_log_folder(folder_name):
    name = str(folder_name).strip()
    if not name:
        return False, "폴더 이름을 입력하세요."

    folders = st.session_state["analysis_log_folders"]
    if name in folders:
        return False, "같은 이름의 프로젝트가 이미 있습니다."

    folders[name] = []
    st.session_state["project_device_settings"][name] = {
        "operating_mode": "Linear",
        "width_um": 1050.0,
        "length_um": 100.0,
        "cox_nf_cm2": 34.5,
    }
    st.session_state["project_workspaces"][name] = (
        default_project_workspace()
    )
    st.session_state["active_log_folder"] = name
    st.session_state["persistent_active_log_id"] = None
    save_projects_state()
    return True, f"'{name}' 프로젝트를 생성했습니다."


def default_project_workspace():
    return {
        "active_file_bytes": None,
        "active_file_name": None,
        "active_file_source": "upload",
        "active_log_id": None,
        "uploader_generation": 0,
        "selected_sheet": None,
    }


def ensure_project_workspace(project_name):
    if not project_name:
        return default_project_workspace()

    workspace = st.session_state["project_workspaces"].setdefault(
        project_name, default_project_workspace()
    )
    defaults = default_project_workspace()
    for key, value in defaults.items():
        workspace.setdefault(key, value)
    return workspace


def save_current_workspace(project_name):
    """Snapshot only the current project's file/log workspace."""
    if not project_name:
        return
    workspace = ensure_project_workspace(project_name)
    workspace["active_file_bytes"] = st.session_state.get(
        "active_file_bytes"
    )
    workspace["active_file_name"] = st.session_state.get(
        "active_file_name"
    )
    workspace["active_file_source"] = st.session_state.get(
        "active_file_source", "upload"
    )
    workspace["active_log_id"] = st.session_state.get(
        "persistent_active_log_id"
    )
    workspace["uploader_generation"] = int(
        st.session_state.get("file_uploader_generation", 0)
    )
    workspace["selected_sheet"] = st.session_state.get(
        "current_selected_sheet"
    )
    save_projects_state()


def load_project_workspace(project_name):
    """Load one project's file/log workspace without touching other projects."""
    workspace = ensure_project_workspace(project_name)

    if workspace.get("active_file_bytes") is not None:
        st.session_state["active_file_bytes"] = workspace[
            "active_file_bytes"
        ]
        st.session_state["active_file_name"] = workspace.get(
            "active_file_name"
        )
    else:
        st.session_state.pop("active_file_bytes", None)
        st.session_state.pop("active_file_name", None)

    st.session_state["active_file_source"] = workspace.get(
        "active_file_source", "upload"
    )
    st.session_state["persistent_active_log_id"] = workspace.get(
        "active_log_id"
    )
    st.session_state["file_uploader_generation"] = int(
        workspace.get("uploader_generation", 0)
    )

    selected_sheet = workspace.get("selected_sheet")
    if selected_sheet:
        st.session_state["restored_sheet"] = selected_sheet
    else:
        st.session_state.pop("restored_sheet", None)


def update_active_project_workspace(
    persist=False,
    **updates,
):
    """Update in-memory workspace; persist only on explicit save actions."""
    project_name = st.session_state.get("active_log_folder")
    if not project_name:
        return

    workspace = ensure_project_workspace(project_name)
    changed = any(
        workspace.get(key) != value
        for key, value in updates.items()
    )
    if changed:
        workspace.update(updates)

    if persist and changed:
        save_projects_state()


def ensure_project_device_settings(project_name):
    if not project_name:
        return {
            "operating_mode": "Linear",
            "width_um": 1050.0,
            "length_um": 100.0,
            "cox_nf_cm2": 34.5,
        }

    settings = st.session_state["project_device_settings"].setdefault(
        project_name,
        {
            "operating_mode": "Linear",
            "width_um": 1050.0,
            "length_um": 100.0,
            "cox_nf_cm2": 34.5,
        },
    )
    settings.setdefault("operating_mode", "Linear")
    settings.setdefault("width_um", 1050.0)
    settings.setdefault("length_um", 100.0)
    settings.setdefault("cox_nf_cm2", 34.5)
    return settings


def save_project_device_widget(project_name, field_name, widget_key):
    if not project_name:
        return
    settings = ensure_project_device_settings(project_name)
    settings[field_name] = st.session_state[widget_key]
    save_projects_state()


def load_log_device_into_project(project_name, record):
    """A saved log updates only its owning project's device information."""
    if not project_name:
        return

    settings = ensure_project_device_settings(project_name)
    mode = record.get("Operating mode")
    if mode in ("Linear", "Saturation"):
        settings["operating_mode"] = mode
    if record.get("Width (μm)") is not None:
        settings["width_um"] = float(record["Width (μm)"])
    if record.get("Length (μm)") is not None:
        settings["length_um"] = float(record["Length (μm)"])
    if record.get("Cox (nF/cm²)") is not None:
        settings["cox_nf_cm2"] = float(record["Cox (nF/cm²)"])
    save_projects_state()


def delete_log_entry(folder_name, entry_id):
    folders = st.session_state["analysis_log_folders"]
    if folder_name not in folders:
        return

    folders[folder_name] = [
        item for item in folders[folder_name]
        if item.get("_log_id") != entry_id
    ]

    if st.session_state.get("persistent_active_log_id") == entry_id:
        st.session_state["persistent_active_log_id"] = None
        st.session_state.pop("active_file_bytes", None)
        st.session_state.pop("active_file_name", None)
        workspace = ensure_project_workspace(folder_name)
        workspace.update(default_project_workspace())

    save_projects_state()


def clear_log_folder(folder_name):
    folders = st.session_state["analysis_log_folders"]
    if folder_name in folders:
        folders[folder_name] = []
        st.session_state["persistent_active_log_id"] = None
        st.session_state.pop("active_file_bytes", None)
        st.session_state.pop("active_file_name", None)
        ensure_project_workspace(folder_name).update(
            default_project_workspace()
        )
        save_projects_state()


def clear_all_logs():
    folders = st.session_state["analysis_log_folders"]
    for name in list(folders.keys()):
        folders[name] = []
    st.session_state["persistent_active_log_id"] = None
    st.session_state.pop("active_file_bytes", None)
    st.session_state.pop("active_file_name", None)
    save_projects_state()


def remove_mobility_point(
    removed_key, source_idx, force_key, selector_keys=()
):
    removed = list(st.session_state.get(removed_key, []))
    source_idx = int(source_idx)
    if source_idx not in removed:
        removed.append(source_idx)
    st.session_state[removed_key] = removed
    st.session_state[force_key] = True
    st.session_state[f"recalc_after_change_{removed_key}"] = True
    # Recalculate defaults only because the underlying active data changed.
    for selector_key in selector_keys:
        if selector_key:
            st.session_state.pop(selector_key, None)


def reset_mobility_points(removed_key, force_key, selector_keys=()):
    st.session_state[removed_key] = []
    st.session_state[force_key] = True
    st.session_state[f"recalc_after_change_{removed_key}"] = True
    for selector_key in selector_keys:
        if selector_key:
            st.session_state.pop(selector_key, None)


def set_state_value(state_key, value):
    st.session_state[state_key] = float(value)


def sci(value, digits=2):
    """HTML scientific notation: xx × 10^xx."""
    if not np.isfinite(value) or value == 0:
        return "N/A" if not np.isfinite(value) else "0"
    exp = int(np.floor(np.log10(abs(value))))
    coef = value / (10 ** exp)
    return f"{coef:.{digits}f} × 10<sup>{exp}</sup>"


def sci_plain(value, digits=2):
    """Excel/plain-text scientific notation."""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return "N/A"

    if not np.isfinite(numeric):
        return "N/A"
    if numeric == 0:
        return "0"

    exp = int(np.floor(np.log10(abs(numeric))))
    coef = numeric / (10 ** exp)
    return f"{coef:.{digits}f} × 10^{exp}"




EXPORT_COLUMNS = [
    "File",
    "Sheet",
    "Drain voltage",
    "Linear or saturation",
    "Gate voltage range",
    "Gate voltage step",
    "On current (A/um)",
    "Off current (A/um)",
    "on-off ratio",
    "Field-effect mobility",
    "threshold voltage (V)",
    "subthreshold swing (mV/dec)",
    "hysteresis (V)",
]


def log_dataframe(folder_name):
    folders = st.session_state["analysis_log_folders"]
    records = folders.get(folder_name, [])
    if not records:
        return pd.DataFrame(columns=EXPORT_COLUMNS)

    rows = []
    for record_order, record in enumerate(records):
        all_sheet_parameters = record.get("_all_sheet_parameters")

        if isinstance(all_sheet_parameters, list) and all_sheet_parameters:
            source_rows = all_sheet_parameters
        else:
            # Backward compatibility for older single-sheet logs.
            source_rows = [{
                "Sheet": record.get("Sheet", ""),
                "Drain voltage (V)": record.get(
                    "Drain voltage (V)", np.nan
                ),
                "Operating mode": record.get(
                    "Operating mode", ""
                ),
                "Gate voltage range": record.get(
                    "Gate voltage range", ""
                ),
                "Gate voltage step (V)": record.get(
                    "Gate voltage step (V)", np.nan
                ),
                "ON current / Width (A/μm)": record.get(
                    "ON current / Width (A/μm)", np.nan
                ),
                "OFF current / Width (A/μm)": record.get(
                    "OFF current / Width (A/μm)", np.nan
                ),
                "ON/OFF ratio": record.get(
                    "ON/OFF ratio", np.nan
                ),
                "Forward mobility (cm²/V·s)": record.get(
                    "Forward mobility (cm²/V·s)", np.nan
                ),
                "Forward Vth (V)": record.get(
                    "Forward Vth (V)", np.nan
                ),
                "Forward SS (mV/dec)": record.get(
                    "Forward SS (mV/dec)", np.nan
                ),
                "Hysteresis (V)": record.get(
                    "Hysteresis (V)", np.nan
                ),
                "_sheet_order": 0,
            }]

        for sheet_row in source_rows:
            rows.append({
                "File": record.get("File", ""),
                "Sheet": sheet_row.get("Sheet", ""),
                "Drain voltage": sheet_row.get(
                    "Drain voltage (V)", np.nan
                ),
                "Linear or saturation": sheet_row.get(
                    "Operating mode", record.get("Operating mode", "")
                ),
                "Gate voltage range": sheet_row.get(
                    "Gate voltage range", ""
                ),
                "Gate voltage step": sheet_row.get(
                    "Gate voltage step (V)", np.nan
                ),
                "On current (A/um)": sci_plain(
                    sheet_row.get(
                        "ON current / Width (A/μm)", np.nan
                    )
                ),
                "Off current (A/um)": sci_plain(
                    sheet_row.get(
                        "OFF current / Width (A/μm)", np.nan
                    )
                ),
                "on-off ratio": sci_plain(
                    sheet_row.get("ON/OFF ratio", np.nan)
                ),
                "Field-effect mobility": sheet_row.get(
                    "Forward mobility (cm²/V·s)", np.nan
                ),
                "threshold voltage (V)": sheet_row.get(
                    "Forward Vth (V)", np.nan
                ),
                "subthreshold swing (mV/dec)": sheet_row.get(
                    "Forward SS (mV/dec)", np.nan
                ),
                "hysteresis (V)": sheet_row.get(
                    "Hysteresis (V)",
                    record.get("Hysteresis (V)", np.nan),
                ),
                "_record_order": record_order,
                "_sheet_order": int(
                    sheet_row.get("_sheet_order", 0)
                ),
            })

    output = pd.DataFrame(rows)
    if output.empty:
        return pd.DataFrame(columns=EXPORT_COLUMNS)

    output = output.sort_values(
        ["_record_order", "_sheet_order"],
        kind="stable",
    ).drop(columns=["_record_order", "_sheet_order"])
    return output.reindex(columns=EXPORT_COLUMNS).reset_index(drop=True)


def autosize_worksheet(ws):
    """Adjust each column width based on cell content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            # multiline content: use longest line
            length = max((len(line) for line in value.splitlines()), default=0)
            max_length = max(max_length, length)

        ws.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 45)


def style_parameter_sheet(ws):
    """Light-green header row with readable alignment."""
    header_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    header_font = Font(bold=True, color="006100")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_worksheet(ws)


def folder_excel_bytes(folder_name):
    df = log_dataframe(folder_name)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Parameters", index=False)
        ws = writer.book["Parameters"]
        style_parameter_sheet(ws)

    output.seek(0)
    return output.getvalue()


def safe_excel_filename(name):
    safe = re.sub(r'[\\/:*?"<>|]+', "_", str(name)).strip()
    return safe or "FET_Analysis_Log"


initialize_log_state()

if st.session_state.get("persistence_warning"):
    st.sidebar.warning(st.session_state.pop("persistence_warning"))


def restore_log_state(record, folder_name=None):
    """Restore the uploaded file and every saved analysis selection."""
    if folder_name:
        st.session_state["active_log_folder"] = folder_name
        load_log_device_into_project(folder_name, record)
    file_bytes = record.get("_file_bytes")
    if not file_bytes:
        st.session_state["restore_error"] = "저장된 원본 파일 데이터가 없습니다."
        return

    st.session_state["restored_log_id"] = record.get("_log_id")
    st.session_state["persistent_active_log_id"] = record.get("_log_id")
    st.session_state["restored_selected_direction"] = record.get(
        "_selected_direction", "Forward"
    )
    st.session_state["restored_sheet_states"] = dict(
        record.get("_sheet_states", {})
    )
    st.session_state["restored_plot_snapshots"] = dict(
        record.get("_plot_snapshots", {})
    )
    st.session_state["restored_state_log_id"] = record.get("_log_id")
    st.session_state["active_file_name"] = record.get("File", "restored.xlsx")
    st.session_state["active_file_bytes"] = file_bytes
    st.session_state["active_file_source"] = "log"
    if folder_name:
        workspace = ensure_project_workspace(folder_name)
        workspace["active_file_bytes"] = file_bytes
        workspace["active_file_name"] = record.get(
            "File", "restored.xlsx"
        )
        workspace["active_file_source"] = "log"
        workspace["active_log_id"] = record.get("_log_id")
    # Change the uploader widget key whenever a log is opened. This destroys
    # the stale uploader instance that still contains a previously uploaded
    # file, so later slider edits/reruns cannot switch back to that file.
    st.session_state["file_uploader_generation"] = (
        int(st.session_state.get("file_uploader_generation", 0)) + 1
    )
    if folder_name:
        ensure_project_workspace(folder_name)[
            "uploader_generation"
        ] = st.session_state["file_uploader_generation"]
        save_projects_state()
    st.session_state["restored_file_name"] = record.get("File", "restored.xlsx")
    st.session_state["restored_file_bytes"] = file_bytes
    st.session_state["restored_sheet"] = record.get("Sheet")

    mode = record.get("Operating mode")
    if mode in ["Linear", "Saturation"]:
        st.session_state["restored_operating_mode"] = mode

    if record.get("Width (μm)") is not None:
        st.session_state["restored_W"] = float(record["Width (μm)"])
    if record.get("Length (μm)") is not None:
        st.session_state["restored_L"] = float(record["Length (μm)"])
    if record.get("Cox (nF/cm²)") is not None:
        st.session_state["restored_Cox_nf"] = float(record["Cox (nF/cm²)"])

    # Applied at the beginning of the next rerun, before these widgets exist.
    st.session_state["device_restore_pending"] = True

    st.session_state["restored_peak_vg_fwd"] = record.get("Forward peak Vg (V)")
    st.session_state["restored_peak_vg_bwd"] = record.get("Backward peak Vg (V)")
    st.session_state["restored_current_vg_fwd"] = record.get("Selected Forward Vg (V)")
    st.session_state["restored_current_vg_bwd"] = record.get("Selected Backward Vg (V)")
    st.session_state["restored_removed_fwd"] = record.get("_removed_fwd_indices", [])
    st.session_state["restored_removed_bwd"] = record.get("_removed_bwd_indices", [])
    st.session_state["restored_on_vg_fwd"] = record.get("_on_vg_fwd")
    st.session_state["restored_on_vg_rev"] = record.get("_on_vg_rev")
    st.session_state["restored_off_vg_fwd"] = record.get("_off_vg_fwd")
    st.session_state["restored_off_vg_rev"] = record.get("_off_vg_rev")
    st.session_state["restored_vth_vg_fwd"] = record.get("_vth_vg_fwd")
    st.session_state["restored_vth_vg_rev"] = record.get("_vth_vg_rev")
    st.session_state["restored_ss_vg_fwd"] = record.get("_ss_vg_fwd")
    st.session_state["restored_ss_vg_rev"] = record.get("_ss_vg_rev")
    st.session_state["restored_ss_range_start_fwd"] = record.get(
        "_ss_range_start_fwd"
    )
    st.session_state["restored_ss_range_end_fwd"] = record.get(
        "_ss_range_end_fwd"
    )
    st.session_state["restored_ss_range_start_rev"] = record.get(
        "_ss_range_start_rev"
    )
    st.session_state["restored_ss_range_end_rev"] = record.get(
        "_ss_range_end_rev"
    )
    st.session_state["restored_mobility_remove_vg_fwd"] = record.get(
        "_mobility_remove_vg_fwd"
    )
    st.session_state["restored_mobility_remove_vg_rev"] = record.get(
        "_mobility_remove_vg_rev"
    )
    st.session_state["restored_log_remove_vg_fwd"] = record.get(
        "_log_remove_vg_fwd"
    )
    st.session_state["restored_log_remove_vg_rev"] = record.get(
        "_log_remove_vg_rev"
    )
    st.session_state["restore_pending"] = True
    save_projects_state()


def auto_restore_last_log():
    """After a browser refresh, automatically reopen the last active saved log."""
    if st.session_state.get("active_file_bytes"):
        return

    active_id = st.session_state.get("persistent_active_log_id")
    if not active_id:
        return

    active_folder = st.session_state.get("active_log_folder")
    records = st.session_state.get(
        "analysis_log_folders", {}
    ).get(active_folder, [])
    for record in records:
        if record.get("_log_id") == active_id:
            restore_log_state(record, active_folder)
            return


def consume_restore_value(key, default=None):
    return st.session_state.pop(key, default)



initialize_log_state()
auto_restore_last_log()

# ============================================================
# Sidebar: Projects first
# ============================================================
initialize_log_state()

st.sidebar.header("Projects")
st.sidebar.caption("프로젝트별로 로그와 Device Information이 독립 저장됩니다.")

project_name_input = st.sidebar.text_input(
    "New project name",
    key="new_project_name_sidebar",
    placeholder="ex) Name",
)

if st.sidebar.button("＋ Create Project", use_container_width=True):
    ok, message = create_log_folder(project_name_input)
    if ok:
        st.sidebar.success(message)
        st.rerun()
    else:
        st.sidebar.warning(message)

project_names = list(st.session_state["analysis_log_folders"].keys())

if project_names:
    # Ensure legacy projects also receive independent defaults.
    for project_name in project_names:
        ensure_project_device_settings(project_name)
        ensure_project_workspace(project_name)

    project_display_names = [f"📁  {name}" for name in project_names]
    selected_display = st.sidebar.radio(
        "Project list",
        project_display_names,
        index=(
            project_names.index(st.session_state["active_log_folder"])
            if st.session_state.get("active_log_folder") in project_names
            else 0
        ),
        key="project_radio_sidebar",
        label_visibility="collapsed",
    )
    active_project = project_names[
        project_display_names.index(selected_display)
    ]

    previous_project = st.session_state.get(
        "workspace_loaded_project"
    )
    if previous_project != active_project:
        if previous_project:
            save_current_workspace(previous_project)

        st.session_state["active_log_folder"] = active_project
        load_project_workspace(active_project)
        st.session_state["workspace_loaded_project"] = active_project
        save_projects_state()

    active_logs = st.session_state["analysis_log_folders"][active_project]
    st.sidebar.caption(
        f"Selected: {active_project} · {len(active_logs)} logs"
    )


    st.sidebar.header("Device Information")

    if active_project:
        project_device = ensure_project_device_settings(active_project)

        mode_key = f"operating_mode_widget__{active_project}"
        width_key = f"width_widget__{active_project}"
        length_key = f"length_widget__{active_project}"
        cox_key = f"cox_widget__{active_project}"

        # Rebuild project-specific widgets from that project's stored values.
        if mode_key not in st.session_state:
            st.session_state[mode_key] = project_device["operating_mode"]
        if width_key not in st.session_state:
            st.session_state[width_key] = float(project_device["width_um"])
        if length_key not in st.session_state:
            st.session_state[length_key] = float(project_device["length_um"])
        if cox_key not in st.session_state:
            st.session_state[cox_key] = float(project_device["cox_nf_cm2"])

        operating_mode = st.sidebar.radio(
            "Operating Mode",
            ["Linear", "Saturation"],
            key=mode_key,
            horizontal=True,
            on_change=save_project_device_widget,
            args=(active_project, "operating_mode", mode_key),
        )
        W = st.sidebar.number_input(
            "Width (μm)",
            min_value=0.000001,
            step=50.0,
            key=width_key,
            on_change=save_project_device_widget,
            args=(active_project, "width_um", width_key),
        )
        L = st.sidebar.number_input(
            "Length (μm)",
            min_value=0.000001,
            step=50.0,
            key=length_key,
            on_change=save_project_device_widget,
            args=(active_project, "length_um", length_key),
        )
        Cox_nf = st.sidebar.number_input(
            "Capacitance (nF/cm⁻²)",
            min_value=0.000001,
            key=cox_key,
            on_change=save_project_device_widget,
            args=(active_project, "cox_nf_cm2", cox_key),
        )

        # Keep model storage synchronized even before a callback fires.
        project_device["operating_mode"] = operating_mode
        project_device["width_um"] = float(W)
        project_device["length_um"] = float(L)
        project_device["cox_nf_cm2"] = float(Cox_nf)
    else:
        operating_mode = "Linear"
        W = 1050.0
        L = 100.0
        Cox_nf = 34.5
        st.sidebar.caption("프로젝트를 선택하면 값을 수정할 수 있습니다.")

    Cox = float(Cox_nf) * 1e-9
    st.sidebar.markdown("---")

    # Populated after an Excel file is available.
    sheet_selector_slot = st.sidebar.container()

    # ============================================================

    if active_logs:
        st.sidebar.download_button(
            "Export Project to Excel",
            data=folder_excel_bytes(active_project),
            file_name=(
                f"{safe_excel_filename(active_project)}_FET_parameters.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key=f"sidebar_export_{active_project}",
        )

        st.sidebar.markdown(
            "<div style='font-size:12px; font-weight:700; "
            "margin:3px 0 2px 0;'>Saved logs</div>",
            unsafe_allow_html=True,
        )
        for log_idx, log_record in enumerate(active_logs, start=1):
            log_name_col, log_delete_col = st.sidebar.columns([8.5, 1])

            saved_at = str(log_record.get("Saved at", ""))
            saved_time = (
                saved_at[11:16] if len(saved_at) >= 16 else ""
            )
            raw_file_name = str(log_record.get("File", ""))
            compact_file_name = (
                raw_file_name
                if len(raw_file_name) <= 24
                else raw_file_name[:21] + "…"
            )
            sheet_count = len(
                log_record.get("_all_sheet_parameters", [])
            )
            log_label = f"📄 {compact_file_name}"
            if sheet_count:
                log_label += f" · {sheet_count} sheets"
            elif log_record.get("Sheet"):
                log_label += f" · {log_record.get('Sheet')}"
            if saved_time:
                log_label += f"  {saved_time}"

            if log_name_col.button(
                log_label,
                key=(
                    f"open_log_{active_project}_"
                    f"{log_record['_log_id']}"
                ),
                use_container_width=True,
                help=(
                    f"{raw_file_name} · "
                    f"{log_record.get('Sheet', '')} · {saved_at}\n"
                    "저장 당시 분석 상태로 열기"
                ),
            ):
                restore_log_state(log_record, active_project)
                save_projects_state()
                st.rerun()

            if log_delete_col.button(
                "×",
                key=(
                    f"sidebar_delete_{active_project}_"
                    f"{log_record['_log_id']}"
                ),
                help="이 로그 삭제",
                use_container_width=True,
            ):
                delete_log_entry(
                    active_project, log_record["_log_id"]
                )
                st.rerun()
    else:
        st.sidebar.info("이 프로젝트에는 저장된 로그가 없습니다.")

    project_action_1, project_action_2 = st.sidebar.columns(2)
    if project_action_1.button(
        "Clear Logs",
        key=f"sidebar_clear_{active_project}",
        use_container_width=True,
    ):
        clear_log_folder(active_project)
        st.rerun()

    if project_action_2.button(
        "Delete Project",
        key=f"sidebar_delete_project_{active_project}",
        use_container_width=True,
    ):
        del st.session_state["analysis_log_folders"][active_project]
        st.session_state["project_device_settings"].pop(
            active_project, None
        )
        st.session_state["project_workspaces"].pop(
            active_project, None
        )
        st.session_state["active_log_folder"] = None
        st.session_state["persistent_active_log_id"] = None
        st.session_state.pop("active_file_bytes", None)
        st.session_state.pop("active_file_name", None)
        save_projects_state()
        st.rerun()
else:
    active_project = None
    st.session_state["active_log_folder"] = None
    st.sidebar.info("먼저 개인 프로젝트를 생성하세요.")

    st.sidebar.markdown("---")
    st.sidebar.header("Device Information")
    operating_mode = "Linear"
    W = 1050.0
    L = 100.0
    Cox_nf = 34.5
    Cox = Cox_nf * 1e-9
    st.sidebar.caption("프로젝트를 선택하면 값을 수정할 수 있습니다.")
    st.sidebar.markdown("---")
    sheet_selector_slot = st.sidebar.container()

st.sidebar.markdown("---")

# Helpers
# ============================================================
def fix_inf(values):
    s = pd.Series(values).replace([np.inf, -np.inf], np.nan)
    return s.ffill().bfill().to_numpy()


def make_card(title, value, color):
    return f"""
    <div style='text-align:left; padding:7px 4px 9px 4px; min-width:0;'>
        <div style='font-size:14px; color:#555; line-height:1.25;
                    min-height:34px; margin-bottom:5px;
                    overflow-wrap:anywhere; word-break:keep-all;'>
            {title}
        </div>
        <div style='font-size:21px; font-weight:750; color:{color};
                    line-height:1.2; min-height:23px;
                    white-space:nowrap; overflow:hidden;
                    text-overflow:ellipsis;'>
            {value}
        </div>
    </div>
    """


def calculate_ss(id_vals, vg_vals):
    """
    기존 앱과 동일한 SS 정의:
    전체 sweep에서 max |d(log10|Id|)/dVg|의 역수.
    """
    id_vals = np.asarray(id_vals, dtype=float)
    vg_vals = np.asarray(vg_vals, dtype=float)

    valid = np.isfinite(id_vals) & np.isfinite(vg_vals)
    if valid.sum() < 3:
        return np.nan

    current = id_vals[valid]
    vg = vg_vals[valid]

    log_id = np.log10(np.abs(current) + 1e-15)
    slope = np.abs(np.gradient(log_id, vg))

    if len(slope) >= 3:
        slope = np.convolve(slope, np.ones(3) / 3.0, mode="same")

    slope = slope[np.isfinite(slope) & (slope > 0)]
    return 1000.0 / np.max(slope) if len(slope) else np.nan



def ss_curve(id_vals, vg_vals):
    """Pointwise SS curve in mV/dec using the same derivative basis as calculate_ss."""
    current = np.asarray(id_vals, dtype=float)
    vg = np.asarray(vg_vals, dtype=float)
    log_id = np.log10(np.abs(current) + 1e-15)
    slope = np.abs(np.gradient(log_id, vg))
    if len(slope) >= 3:
        slope = np.convolve(slope, np.ones(3) / 3.0, mode="same")
    out = np.full(len(slope), np.nan, dtype=float)
    valid = np.isfinite(slope) & (slope > 0)
    out[valid] = 1000.0 / slope[valid]
    return out


def vth_at_index(vg_vals, id_vals, gm_vals, idx, mode):
    vg = np.asarray(vg_vals, dtype=float)
    current = np.asarray(id_vals, dtype=float)
    gm = np.asarray(gm_vals, dtype=float)
    idx = int(np.clip(idx, 0, len(vg) - 1))
    gm_value = gm[idx]
    if not np.isfinite(gm_value) or abs(gm_value) <= np.finfo(float).eps:
        return np.nan
    numerator = np.sqrt(abs(current[idx])) if mode == "Saturation" else current[idx]
    return float(vg[idx] - numerator / gm_value)


def split_sweep(df):
    work = df.copy().reset_index(drop=True)
    work["__source_index"] = np.arange(len(work), dtype=int)
    work["GateV"] = pd.to_numeric(work["GateV"], errors="coerce")
    work["DrainI"] = pd.to_numeric(work["DrainI"], errors="coerce")
    work = work.dropna(subset=["GateV", "DrainI"]).reset_index(drop=True)

    vg = work["GateV"]
    if abs(vg.max() - vg.iloc[0]) > abs(vg.min() - vg.iloc[0]):
        turning = int(vg.idxmax())
    else:
        turning = int(vg.idxmin())

    fwd = work.iloc[:turning + 1].reset_index(drop=True)
    bwd = work.iloc[turning:].reset_index(drop=True)
    return fwd, bwd


def calc_curves(vg, current, mode, w, l, cox, vd):
    vg = np.asarray(vg, dtype=float)
    current = np.asarray(current, dtype=float)

    if len(vg) < 3:
        nan = np.full(len(vg), np.nan)
        return nan, nan

    if mode == "Linear":
        gm = fix_inf(np.gradient(current, vg))
        mobility = np.abs(gm) * l / (w * cox * abs(vd))
    else:
        sqrt_id = np.sqrt(np.abs(current))
        gm = fix_inf(np.gradient(sqrt_id, vg))
        mobility = (2.0 * l / (w * cox)) * gm**2

    return gm, mobility



@st.cache_data(show_spinner=False, max_entries=256)
def cached_calc_curves(
    vg_tuple,
    current_tuple,
    mode,
    w,
    l,
    cox,
    vd,
):
    return calc_curves(
        np.asarray(vg_tuple, dtype=float),
        np.asarray(current_tuple, dtype=float),
        mode,
        float(w),
        float(l),
        float(cox),
        float(vd),
    )

def auto_peak_index(mobility):
    """Detect the point with the largest abnormal local mobility change.

    The detection score uses only local change:
    - deviation from the line connecting the neighboring points
    - abrupt changes on both sides of the center point

    Mobility magnitude itself is not included in the score.
    """
    values = np.asarray(mobility, dtype=float)
    n = len(values)

    if n == 0:
        return 0

    finite_mask = np.isfinite(values)
    if not finite_mask.any():
        return 0

    if n < 5:
        valid_indices = np.where(finite_mask)[0]
        if len(valid_indices) == 1:
            return int(valid_indices[0])

        work_small = values.copy()
        indices_small = np.arange(n)
        work_small[~finite_mask] = np.interp(
            indices_small[~finite_mask],
            indices_small[finite_mask],
            work_small[finite_mask],
        )
        differences = np.abs(np.gradient(work_small))
        differences[~finite_mask] = -np.inf
        return int(np.argmax(differences))

    work = values.copy()
    valid = np.isfinite(work)

    if valid.sum() < 3:
        valid_indices = np.where(valid)[0]
        return int(valid_indices[0])

    point_indices = np.arange(n)
    work[~valid] = np.interp(
        point_indices[~valid],
        point_indices[valid],
        work[valid],
    )

    # Difference from the local straight line through neighboring points.
    neighbor_average = 0.5 * (
        work[:-2] + work[2:]
    )
    curvature = np.abs(
        work[1:-1] - neighbor_average
    )

    # An isolated spike changes sharply on both sides.
    left_change = np.abs(
        work[1:-1] - work[:-2]
    )
    right_change = np.abs(
        work[2:] - work[1:-1]
    )
    two_sided_change = np.minimum(
        left_change,
        right_change,
    )

    def robust_positive_score(array):
        array = np.asarray(array, dtype=float)
        median = np.nanmedian(array)
        mad = np.nanmedian(
            np.abs(array - median)
        )
        scale = max(
            1.4826 * mad,
            np.finfo(float).eps,
        )
        return np.maximum(
            (array - median) / scale,
            0.0,
        )

    local_change_score = (
        robust_positive_score(curvature)
        + robust_positive_score(two_sided_change)
    )

    # Avoid unreliable points immediately beside the boundaries.
    local_change_score[:1] = -np.inf
    local_change_score[-1:] = -np.inf

    finite_scores = np.isfinite(local_change_score)
    if not finite_scores.any():
        differences = np.abs(np.gradient(work))
        differences[0] = -np.inf
        differences[-1] = -np.inf
        return int(np.nanargmax(differences))

    if np.nanmax(local_change_score[finite_scores]) <= 0:
        differences = np.abs(np.gradient(work))
        differences[0] = -np.inf
        differences[-1] = -np.inf
        return int(np.nanargmax(differences))

    return int(np.nanargmax(local_change_score)) + 1


def parameter_values(
    vg_fwd, id_fwd, gm_fwd, mu_fwd, idx_f,
    vg_bwd, id_bwd, gm_bwd, mu_bwd, idx_b,
    mode, width, on_current_override=None, off_current_override=None
):
    vg_fwd = pd.Series(vg_fwd).reset_index(drop=True)
    id_fwd = pd.Series(id_fwd).reset_index(drop=True)
    vg_bwd = pd.Series(vg_bwd).reset_index(drop=True)
    id_bwd = pd.Series(id_bwd).reset_index(drop=True)

    def safe_vth(vg, current, gm, idx):
        if len(vg) == 0 or idx >= len(vg):
            return np.nan

        gm_value = gm[idx]
        if not np.isfinite(gm_value) or abs(gm_value) <= np.finfo(float).eps:
            return np.nan

        numerator = (
            np.sqrt(abs(current.iloc[idx]))
            if mode == "Saturation"
            else current.iloc[idx]
        )
        return float(vg.iloc[idx] - numerator / gm_value)

    vth_f = safe_vth(vg_fwd, id_fwd, gm_fwd, idx_f)
    vth_b = safe_vth(vg_bwd, id_bwd, gm_bwd, idx_b)

    full_current = np.concatenate([
        np.asarray(id_fwd, dtype=float),
        np.asarray(id_bwd, dtype=float)[1:],
    ])

    finite_abs = np.abs(full_current[np.isfinite(full_current)])
    positive_abs = finite_abs[finite_abs > 0]

    default_on_current = float(np.max(finite_abs)) if len(finite_abs) else np.nan
    default_off_current = float(np.min(positive_abs)) if len(positive_abs) else np.nan

    on_current = (
        float(on_current_override)
        if on_current_override is not None and np.isfinite(on_current_override)
        else default_on_current
    )
    off_current = (
        float(off_current_override)
        if off_current_override is not None and np.isfinite(off_current_override)
        else default_off_current
    )

    return {
        "mu_fwd": float(mu_fwd[idx_f]) if len(mu_fwd) else np.nan,
        "mu_bwd": float(mu_bwd[idx_b]) if len(mu_bwd) else np.nan,
        "vth_fwd": vth_f,
        "vth_bwd": vth_b,
        "peak_vg_fwd": float(vg_fwd.iloc[idx_f]) if len(vg_fwd) else np.nan,
        "peak_vg_bwd": float(vg_bwd.iloc[idx_b]) if len(vg_bwd) else np.nan,
        "ss_fwd": calculate_ss(id_fwd, vg_fwd),
        "ss_bwd": calculate_ss(id_bwd, vg_bwd),
        "hysteresis": (
            abs(vth_f - vth_b)
            if np.isfinite(vth_f) and np.isfinite(vth_b)
            else np.nan
        ),
        "onoff": (
            on_current / off_current
            if np.isfinite(off_current) and off_current > 0
            else np.nan
        ),
        "on_density": on_current / width if np.isfinite(on_current) else np.nan,
        "off_density": off_current / width if np.isfinite(off_current) else np.nan,
    }


def current_density_at_vg(active_df, selected_vg, width):
    """선택한 실제 Vg에서 |DrainI|/Width를 반환."""
    idx = int((active_df["GateV"] - float(selected_vg)).abs().idxmin())
    row = active_df.iloc[idx]
    density = abs(float(row["DrainI_active"])) / float(width)
    return idx, row, density


def state_keys(file_id, sheet_name, mode):
    stem = f"{file_id}_{sheet_name}_{mode}"
    return {
        "removed_fwd": f"removed_fwd_{stem}",
        "removed_bwd": f"removed_bwd_{stem}",
        "remove_slider_fwd": f"remove_slider_fwd_{stem}",
        "remove_slider_bwd": f"remove_slider_bwd_{stem}",
        "log_remove_slider_fwd": f"log_remove_slider_fwd_{stem}",
        "log_remove_slider_bwd": f"log_remove_slider_bwd_{stem}",
        "peak_slider_fwd": f"peak_slider_fwd_{stem}",
        "peak_slider_bwd": f"peak_slider_bwd_{stem}",
        "ss_slider_fwd": f"ss_slider_fwd_{stem}",
        "ss_slider_bwd": f"ss_slider_bwd_{stem}",
        "ss_range_start_fwd": f"ss_range_start_fwd_{stem}",
        "ss_range_start_bwd": f"ss_range_start_bwd_{stem}",
        "ss_range_end_fwd": f"ss_range_end_fwd_{stem}",
        "ss_range_end_bwd": f"ss_range_end_bwd_{stem}",
        "force_auto_peak_fwd": f"force_auto_peak_fwd_{stem}",
        "force_auto_peak_bwd": f"force_auto_peak_bwd_{stem}",
        "current_slider_fwd": f"current_slider_fwd_{stem}",
        "current_slider_bwd": f"current_slider_bwd_{stem}",
    }


def initialize_removal_state(keys):
    if keys["removed_fwd"] not in st.session_state:
        st.session_state[keys["removed_fwd"]] = []
    if keys["removed_bwd"] not in st.session_state:
        st.session_state[keys["removed_bwd"]] = []
    if keys["force_auto_peak_fwd"] not in st.session_state:
        st.session_state[keys["force_auto_peak_fwd"]] = False
    if keys["force_auto_peak_bwd"] not in st.session_state:
        st.session_state[keys["force_auto_peak_bwd"]] = False


def build_active_sweep(part, removed_source_indices):
    """수동 제거된 원래 행을 제외하고 나머지 raw 데이터를 그대로 사용."""
    removed = set(int(v) for v in removed_source_indices)
    active = part[~part["__source_index"].isin(removed)].copy().reset_index(drop=True)
    active["DrainI_active"] = active["DrainI"].to_numpy()
    return active



def normalize_excel_label(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def standardize_measurement_columns(df):
    aliases = {
        "gatev": "GateV",
        "gatevoltage": "GateV",
        "vg": "GateV",
        "draini": "DrainI",
        "draincurrent": "DrainI",
        "id": "DrainI",
        "drainv": "DrainV",
        "drainvoltage": "DrainV",
        "vd": "DrainV",
        "vds": "DrainV",
        "gatei": "GateI",
        "gatecurrent": "GateI",
        "ig": "GateI",
    }
    rename_map = {}
    for column in df.columns:
        normalized = normalize_excel_label(column)
        if normalized in aliases:
            rename_map[column] = aliases[normalized]
    return df.rename(columns=rename_map)


def extract_drain_v_from_settings(xls, selected_sheet=None):
    """Read Drain V from the selected sheet's condition region in Settings.

    Procedure:
    1) Find the selected sheet's block in Settings.
    2) In column A of that block, find the row labeled 'Name'.
    3) In that same row, find the column labeled 'DrainV'.
    4) In column A, find the row labeled 'Start/Level'.
    5) Return the cell at the intersection of the DrainV column and
       Start/Level row.
    """
    settings_sheet = next(
        (
            name for name in xls.sheet_names
            if name.strip().lower() == "settings"
        ),
        None,
    )
    if settings_sheet is None:
        return None

    try:
        raw = xls.parse(settings_sheet, header=None)
    except Exception:
        return None

    rows, cols = raw.shape
    selected_norm = normalize_excel_label(selected_sheet or "")

    # Find selected sheet anchors anywhere in Settings.
    anchors = []
    if selected_norm:
        for r in range(rows):
            for c in range(cols):
                if normalize_excel_label(raw.iat[r, c]) == selected_norm:
                    anchors.append((r, c))

    # If the sheet name is not explicitly written, use the whole sheet.
    if not anchors:
        anchors = [(0, 0)]

    for anchor_r, anchor_c in anchors:
        # Build a reasonable condition-region window around the anchor.
        r0 = max(0, anchor_r - 3)
        r1 = min(rows, anchor_r + 80)
        c0 = 0
        c1 = cols

        name_row = None
        start_level_row = None

        # User specified that labels are in column A.
        for r in range(r0, r1):
            a_token = normalize_excel_label(raw.iat[r, 0])
            if a_token == "name":
                name_row = r
            elif a_token in {"startlevel", "start", "level"}:
                start_level_row = r

        if name_row is None or start_level_row is None:
            continue

        drainv_col = None
        for c in range(c0, c1):
            token = normalize_excel_label(raw.iat[name_row, c])
            if token in {"drainv", "drainvoltage", "vd", "vds"}:
                drainv_col = c
                break

        if drainv_col is None:
            continue

        value = pd.to_numeric(
            pd.Series([raw.iat[start_level_row, drainv_col]]),
            errors="coerce",
        ).iloc[0]
        if pd.notna(value):
            return float(value)

    return None


def analyze_sheet(df, file_id, sheet_name):
    if W <= 0 or L <= 0 or Cox <= 0:
        raise ValueError("Width, Length, Capacitance는 모두 0보다 커야 합니다.")

    vd_values = pd.to_numeric(df["DrainV"], errors="coerce").dropna()
    if vd_values.empty:
        raise ValueError("DrainV 값이 없습니다.")

    vd = float(vd_values.iloc[0])
    if operating_mode == "Linear" and abs(vd) <= np.finfo(float).eps:
        raise ValueError("Linear mode에서는 DrainV가 0이 아니어야 합니다.")

    fwd_raw, bwd_raw = split_sweep(df)
    keys = state_keys(file_id, sheet_name, operating_mode)
    initialize_removal_state(keys)

    fwd = build_active_sweep(fwd_raw, st.session_state[keys["removed_fwd"]])
    bwd = build_active_sweep(bwd_raw, st.session_state[keys["removed_bwd"]])

    if len(fwd) < 3 or len(bwd) < 3:
        raise ValueError("점 제거 후 각 sweep에 최소 3개의 데이터가 필요합니다.")

    gm_fwd, mu_fwd = cached_calc_curves(
        tuple(
            pd.to_numeric(
                fwd["GateV"], errors="coerce"
            ).to_numpy(dtype=float)
        ),
        tuple(
            pd.to_numeric(
                fwd["DrainI_active"], errors="coerce"
            ).to_numpy(dtype=float)
        ),
        operating_mode, W, L, Cox, vd,
    )
    gm_bwd, mu_bwd = cached_calc_curves(
        tuple(
            pd.to_numeric(
                bwd["GateV"], errors="coerce"
            ).to_numpy(dtype=float)
        ),
        tuple(
            pd.to_numeric(
                bwd["DrainI_active"], errors="coerce"
            ).to_numpy(dtype=float)
        ),
        operating_mode, W, L, Cox, vd,
    )

    # Peak Elimination target: strongest isolated local anomaly.
    auto_idx_f = auto_peak_index(mu_fwd)
    auto_idx_b = auto_peak_index(mu_bwd)

    # Mobility Auto Set/default: maximum finite mobility in the current
    # active data after all eliminated points have been removed.
    finite_mu_f = np.where(np.isfinite(mu_fwd), mu_fwd, -np.inf)
    finite_mu_b = np.where(np.isfinite(mu_bwd), mu_bwd, -np.inf)
    mobility_max_idx_f = (
        int(np.argmax(finite_mu_f))
        if np.any(np.isfinite(mu_fwd)) else 0
    )
    mobility_max_idx_b = (
        int(np.argmax(finite_mu_b))
        if np.any(np.isfinite(mu_bwd)) else 0
    )

    # First run or post-elimination rerun resets mobility selection to
    # the current maximum mobility point.
    if (
        keys["peak_slider_fwd"] not in st.session_state
        or st.session_state.get(keys["force_auto_peak_fwd"], False)
    ):
        st.session_state[keys["peak_slider_fwd"]] = float(
            fwd["GateV"].iloc[mobility_max_idx_f]
        )
        st.session_state[keys["force_auto_peak_fwd"]] = False

    if (
        keys["peak_slider_bwd"] not in st.session_state
        or st.session_state.get(keys["force_auto_peak_bwd"], False)
    ):
        st.session_state[keys["peak_slider_bwd"]] = float(
            bwd["GateV"].iloc[mobility_max_idx_b]
        )
        st.session_state[keys["force_auto_peak_bwd"]] = False

    peak_target_f = float(st.session_state[keys["peak_slider_fwd"]])
    peak_target_b = float(st.session_state[keys["peak_slider_bwd"]])

    idx_f = int((fwd["GateV"] - peak_target_f).abs().idxmin())
    idx_b = int((bwd["GateV"] - peak_target_b).abs().idxmin())

    # 실제 남아 있는 Vg에 snap
    st.session_state[keys["peak_slider_fwd"]] = float(fwd["GateV"].iloc[idx_f])
    st.session_state[keys["peak_slider_bwd"]] = float(bwd["GateV"].iloc[idx_b])

    params = parameter_values(
        fwd["GateV"], fwd["DrainI_active"], gm_fwd, mu_fwd, idx_f,
        bwd["GateV"], bwd["DrainI_active"], gm_bwd, mu_bwd, idx_b,
        operating_mode, W,
    )

    return {
        "vd": vd,
        "keys": keys,
        "fwd_raw": fwd_raw,
        "bwd_raw": bwd_raw,
        "fwd": fwd,
        "bwd": bwd,
        "gm_fwd": gm_fwd,
        "gm_bwd": gm_bwd,
        "mu_fwd": mu_fwd,
        "mu_bwd": mu_bwd,
        "idx_f": idx_f,
        "idx_b": idx_b,
        "auto_idx_f": auto_idx_f,
        "auto_idx_b": auto_idx_b,
        "mobility_max_idx_f": mobility_max_idx_f,
        "mobility_max_idx_b": mobility_max_idx_b,
        "params": params,
    }


def nearest_row_by_vg(active_df, selected_vg):
    idx = int((active_df["GateV"] - float(selected_vg)).abs().idxmin())
    return idx, active_df.iloc[idx]


def snapped_slider_point(active_df, state_key):
    """Snap a persistent slider value once and return that exact data point.

    The snapped value is written back to session_state so the slider, plotted
    arrow and remove callback always refer to the same measured Vg row.
    """
    values = sorted_unique_vg(active_df)
    if len(values) == 0:
        return None, None

    requested = float(
        st.session_state.get(state_key, values[0])
    )
    nearest_value_idx = int(
        np.argmin(np.abs(values - requested))
    )
    snapped_vg = float(values[nearest_value_idx])
    st.session_state[state_key] = snapped_vg

    row_idx = int(
        (
            pd.to_numeric(
                active_df["GateV"], errors="coerce"
            )
            - snapped_vg
        ).abs().idxmin()
    )
    return row_idx, active_df.iloc[row_idx]




def sorted_unique_vg(active_df):
    """현재 sweep에 남아 있는 실제 Vg 값 목록."""
    values = pd.to_numeric(active_df["GateV"], errors="coerce").dropna().unique()
    return np.sort(values.astype(float))


def initialize_slider_in_range(key, active_df, default_value):
    """위젯이 생성되기 전에만 session_state를 초기화한다."""
    values = sorted_unique_vg(active_df)
    if len(values) == 0:
        return

    if key not in st.session_state:
        nearest = int(np.argmin(np.abs(values - float(default_value))))
        st.session_state[key] = float(values[nearest])
        return

    current = float(st.session_state[key])
    nearest = int(np.argmin(np.abs(values - current)))
    nearest_value = float(values[nearest])

    # 현재 값이 삭제된 Vg라면 위젯 생성 전에 가장 가까운 남은 Vg로 보정
    if not np.isclose(current, nearest_value):
        st.session_state[key] = nearest_value


def step_discrete_slider(
    persistent_key,
    widget_key,
    active_df,
    direction,
    input_widget_key=None,
):
    """Move one measured-Vg step and persist it independently of widget life."""
    values = sorted_unique_vg(active_df)
    if len(values) == 0:
        return

    current = float(st.session_state.get(persistent_key, values[0]))
    nearest = int(np.argmin(np.abs(values - current)))
    target = int(np.clip(nearest + int(direction), 0, len(values) - 1))
    target_value = float(values[target])

    st.session_state[persistent_key] = target_value
    st.session_state[widget_key] = target_value
    if input_widget_key:
        st.session_state[input_widget_key] = target_value


def sync_persistent_from_slider(
    widget_key,
    persistent_key,
    input_widget_key,
):
    value = float(st.session_state[widget_key])
    st.session_state[persistent_key] = value
    st.session_state[input_widget_key] = value


def sync_persistent_from_numeric(
    input_widget_key,
    widget_key,
    persistent_key,
    active_df,
):
    values = sorted_unique_vg(active_df)
    if len(values) == 0:
        return

    requested = float(st.session_state[input_widget_key])
    nearest = int(np.argmin(np.abs(values - requested)))
    snapped = float(values[nearest])

    st.session_state[persistent_key] = snapped
    st.session_state[widget_key] = snapped
    st.session_state[input_widget_key] = snapped


def sync_persistent_number(widget_key, persistent_key):
    st.session_state[persistent_key] = float(
        st.session_state[widget_key]
    )


def render_persistent_number_input(
    container,
    label,
    persistent_key,
    widget_key,
    default_value,
    step=0.1,
    fmt="%.3f",
):
    """Render a number input whose value survives when the widget disappears."""
    if persistent_key not in st.session_state:
        st.session_state[persistent_key] = float(default_value)

    # A widget key may have been cleaned up while its direction was hidden.
    # Always rebuild it from the persistent model value.
    st.session_state[widget_key] = float(
        st.session_state[persistent_key]
    )
    container.number_input(
        label,
        key=widget_key,
        step=float(step),
        format=fmt,
        on_change=sync_persistent_number,
        args=(widget_key, persistent_key),
    )
    return float(st.session_state[persistent_key])


def render_discrete_vg_control(
    title,
    slider_label,
    state_key,
    active_df,
    default_value,
    button_prefix,
    parent=None,
):
    """Measured-Vg slider with direction-persistent model state."""
    initialize_slider_in_range(state_key, active_df, default_value)

    values = sorted_unique_vg(active_df)
    if len(values) == 0:
        return np.nan

    options = [float(v) for v in values]
    ui = parent if parent is not None else st.sidebar

    # state_key is persistent model state. These two are disposable widget keys.
    slider_widget_key = f"{button_prefix}_slider_widget"
    input_widget_key = f"{button_prefix}_numeric_widget"

    persistent_value = float(st.session_state[state_key])
    nearest = int(np.argmin(np.abs(np.asarray(options) - persistent_value)))
    persistent_value = float(options[nearest])
    st.session_state[state_key] = persistent_value

    # Recreate disposable widget state from the persistent value every time the
    # direction becomes visible again. This prevents stale widget cleanup from
    # changing Forward/Reverse selections.
    st.session_state[slider_widget_key] = persistent_value
    st.session_state[input_widget_key] = persistent_value

    minus_col, slider_col, plus_col = ui.columns([1, 5, 1])

    minus_col.button(
        "−",
        key=f"{button_prefix}_minus",
        use_container_width=True,
        on_click=step_discrete_slider,
        args=(
            state_key,
            slider_widget_key,
            active_df,
            -1,
            input_widget_key,
        ),
    )

    slider_col.select_slider(
        slider_label,
        options=options,
        key=slider_widget_key,
        label_visibility="collapsed",
        format_func=lambda value: f"{value:.2f}",
        on_change=sync_persistent_from_slider,
        args=(
            slider_widget_key,
            state_key,
            input_widget_key,
        ),
    )

    plus_col.button(
        "+",
        key=f"{button_prefix}_plus",
        use_container_width=True,
        on_click=step_discrete_slider,
        args=(
            state_key,
            slider_widget_key,
            active_df,
            +1,
            input_widget_key,
        ),
    )

    ui.number_input(
        "Vg input (V)",
        key=input_widget_key,
        min_value=float(options[0]),
        max_value=float(options[-1]),
        step=(
            float(np.median(np.diff(options)))
            if len(options) > 1
            else 0.1
        ),
        format="%.3f",
        label_visibility="collapsed",
        on_change=sync_persistent_from_numeric,
        args=(
            input_widget_key,
            slider_widget_key,
            state_key,
            active_df,
        ),
    )

    snapped_index, snapped_row = snapped_slider_point(
        active_df, state_key
    )
    if snapped_row is None:
        return np.nan
    return float(snapped_row["GateV"])



# ============================================================
# Upload
# ============================================================

if st.session_state.get("restore_error"):
    st.error(st.session_state.pop("restore_error"))

current_project = st.session_state.get("active_log_folder")
project_info_col, project_save_col, project_add_col = st.columns(
    [4.8, 1.0, 1.4], gap="small"
)
project_info_col.markdown(
    f"<div style='font-size:24px; font-weight:850; padding-top:1px;'>"
    f"Project: <b>{current_project or 'None'}</b></div>",
    unsafe_allow_html=True,
)
if project_save_col.button(
    "Save",
    key="save_active_log_top_global",
    use_container_width=True,
    disabled=(
        st.session_state.get("persistent_active_log_id") is None
    ),
):
    st.session_state["save_current_requested"] = True

if project_add_col.button(
    "Add to Project",
    key="add_project_top_global",
    use_container_width=True,
    disabled=current_project is None,
):
    st.session_state["add_project_requested"] = True

current_project = st.session_state.get(
    "active_log_folder"
)
uploader_generation = int(
    st.session_state.get("file_uploader_generation", 0)
)
uploader_value = st.file_uploader(
    "측정된 엑셀 파일을 업로드하세요",
    type=["xlsx", "xls"],
    key=(
        f"measurement_file_uploader_{current_project}_"
        f"{uploader_generation}"
    ),
)
main_content = st.container()

# One authoritative source at a time.
#
# Opening a log increments file_uploader_generation, so an old uploaded file
# cannot survive into the log view. A file selected in the CURRENT uploader
# generation is therefore always an explicit new upload and takes priority.
active_source = st.session_state.get("active_file_source", "upload")
uploaded_file = None

if uploader_value is not None:
    try:
        uploader_value.seek(0)
        new_upload_bytes = uploader_value.read()
        uploader_value.seek(0)
    except Exception:
        new_upload_bytes = None

    if new_upload_bytes:
        st.session_state["active_file_source"] = "upload"
        st.session_state["active_file_bytes"] = new_upload_bytes
        st.session_state["active_file_name"] = getattr(
            uploader_value, "name", "uploaded.xlsx"
        )
        # A newly uploaded file is not the previously opened saved log.
        # Disconnect the old log ID so Save cannot overwrite it accidentally.
        st.session_state["persistent_active_log_id"] = None
        st.session_state.pop("restored_log_id", None)
        update_active_project_workspace(
            persist=True,
            active_file_bytes=new_upload_bytes,
            active_file_name=getattr(
                uploader_value, "name", "uploaded.xlsx"
            ),
            active_file_source="upload",
            active_log_id=None,
            uploader_generation=uploader_generation,
        )
        uploaded_file = uploader_value

elif active_source == "log" and st.session_state.get("active_file_bytes"):
    restored_buffer = io.BytesIO(st.session_state["active_file_bytes"])
    restored_buffer.name = st.session_state.get(
        "active_file_name", "restored.xlsx"
    )
    uploaded_file = restored_buffer

elif st.session_state.get("active_file_bytes"):
    fallback_buffer = io.BytesIO(st.session_state["active_file_bytes"])
    fallback_buffer.name = st.session_state.get(
        "active_file_name", "restored.xlsx"
    )
    uploaded_file = fallback_buffer

with main_content:
    if uploaded_file:
        file_name = getattr(uploaded_file, "name", "uploaded.xlsx")
        file_size = getattr(uploaded_file, "size", None)
        if file_size is None:
            try:
                current_pos = uploaded_file.tell()
                uploaded_file.seek(0, 2)
                file_size = uploaded_file.tell()
                uploaded_file.seek(current_pos)
            except Exception:
                try:
                    file_size = len(uploaded_file.getvalue())
                except Exception:
                    file_size = 0
        file_id = f"{file_name}_{file_size}"

        try:
            uploaded_file.seek(0)
            active_excel_bytes = uploaded_file.read()
            uploaded_file.seek(0)
        except Exception:
            active_excel_bytes = st.session_state.get(
                "active_file_bytes", b""
            )

        sheet_names = cached_excel_sheet_names(
            active_excel_bytes
        )
        target_sheets = [
            sheet_name for sheet_name in sheet_names
            if sheet_name.strip().lower()
            not in {"calc", "settings"}
        ]

        if not target_sheets:
            st.error("Calc와 Settings를 제외한 분석 가능한 시트가 없습니다.")
            st.stop()

        restored_sheet = st.session_state.get("restored_sheet")
        default_sheet = (
            restored_sheet if restored_sheet in target_sheets
            else ("Data" if "Data" in target_sheets else target_sheets[0])
        )
        with sheet_selector_slot:
            st.markdown(
                "<div style='font-size:18px;font-weight:750;margin:0 0 5px 0;'>"
                "Select Data Sheet</div>",
                unsafe_allow_html=True,
            )
            selected_sheet = st.selectbox(
                "Select Data Sheet",
                target_sheets,
                index=target_sheets.index(default_sheet),
                key=(
                    f"sheet_selector_{current_project}_{file_id}"
                ),
                label_visibility="collapsed",
            )
            st.session_state["current_selected_sheet"] = selected_sheet
            update_active_project_workspace(
                selected_sheet=selected_sheet
            )
            st.markdown("---")

        # ========================================================
        # Average mode
        # ========================================================
        if False:
            rows = []

            for sheet in target_sheets:
                df_sheet = cached_read_excel_sheet(
                    active_excel_bytes,
                    sheet,
                ).copy()
                if {"GateV", "DrainI", "DrainV"}.issubset(df_sheet.columns):
                    try:
                        result = analyze_sheet(df_sheet, file_id, sheet)
                        rows.append({"Sheet": sheet, **result["params"]})
                    except Exception:
                        pass

            if not rows:
                st.error("유효한 시트가 없습니다.")
                st.stop()

            stats = pd.DataFrame(rows)
            st.markdown(f"### 📊 Statistics ({operating_mode})")

            p = {
                key: stats[key].mean()
                for key in [
                    "mu_fwd", "mu_bwd", "vth_fwd", "vth_bwd",
                    "ss_fwd", "ss_bwd", "hysteresis",
                    "onoff", "on_density", "off_density",
                ]
            }

            st.markdown("<h4 style='color:#6FADCF;'>Forward Sweep Parameters</h4>", unsafe_allow_html=True)
            f1, f2, f3 = st.columns(3)
            f1.markdown(make_card("Peak Mobility", f"{p['mu_fwd']:.2f} cm²/V·s", "#2E60AB"), unsafe_allow_html=True)
            f2.markdown(make_card("Threshold Voltage", f"{p['vth_fwd']:.2f} V", "#A23B72"), unsafe_allow_html=True)
            f3.markdown(make_card("SS", f"{p['ss_fwd']:.1f} mV/dec", "#18A558"), unsafe_allow_html=True)

            st.markdown("<h4 style='color:#F05650;'>Backward Sweep Parameters</h4>", unsafe_allow_html=True)
            b1, b2, b3 = st.columns(3)
            b1.markdown(make_card("Peak Mobility", f"{p['mu_bwd']:.2f} cm²/V·s", "#2E60AB"), unsafe_allow_html=True)
            b2.markdown(make_card("Threshold Voltage", f"{p['vth_bwd']:.2f} V", "#A23B72"), unsafe_allow_html=True)
            b3.markdown(make_card("SS", f"{p['ss_bwd']:.1f} mV/dec", "#18A558"), unsafe_allow_html=True)

            st.markdown("<h4>Overall Device Parameters</h4>", unsafe_allow_html=True)
            o1, o2, o3, o4 = st.columns(4)
            o1.markdown(make_card("On/Off Ratio", sci(p["onoff"]), "#5B5F97"), unsafe_allow_html=True)
            o2.markdown(make_card("ON Current / Width", f"{sci(p['on_density'])} A/μm", "#5B5F97"), unsafe_allow_html=True)
            o3.markdown(make_card("OFF Current / Width", f"{sci(p['off_density'])} A/μm", "#5B5F97"), unsafe_allow_html=True)
            o4.markdown(make_card("Hysteresis", f"{p['hysteresis']:.2f} V", "#5B5F97"), unsafe_allow_html=True)

        # ========================================================
        # Single-sheet mode
        # ========================================================
        else:
            df = standardize_measurement_columns(
                cached_read_excel_sheet(
                    active_excel_bytes,
                    selected_sheet,
                ).copy()
            )

            if not {"GateV", "DrainI"}.issubset(df.columns):
                st.error("GateV와 DrainI 컬럼이 필요합니다.")
                st.stop()

            if "DrainV" not in df.columns:
                settings_drain_v = cached_settings_drain_v(
                    active_excel_bytes,
                    selected_sheet,
                )
                if settings_drain_v is None:
                    st.error(
                        "선택한 시트에 DrainV가 없고 Settings 시트에서도 "
                        "Drain V 값을 찾지 못했습니다."
                    )
                    st.stop()
                df["DrainV"] = float(settings_drain_v)
                st.caption(
                    f"Drain V = {settings_drain_v:g} V "
                    f"(Settings 시트에서 불러옴)"
                )

            # Apply saved state before analysis so removed points and selected peaks
            # are reflected immediately in the reconstructed curves.
            pre_keys = state_keys(file_id, selected_sheet, operating_mode)
            initialize_removal_state(pre_keys)

            restored_sheet_states = st.session_state.get(
                "restored_sheet_states", {}
            )
            restored_log_id = st.session_state.get(
                "restored_state_log_id",
                st.session_state.get("restored_log_id", "legacy"),
            )
            sheet_restore_token = (
                f"sheet_restore_applied_{restored_log_id}_"
                f"{file_id}_{selected_sheet}_{operating_mode}"
            )

            saved_sheet_state = (
                restored_sheet_states.get(selected_sheet)
                if isinstance(restored_sheet_states, dict)
                else None
            )

            if (
                isinstance(saved_sheet_state, dict)
                and not st.session_state.get(
                    sheet_restore_token, False
                )
            ):
                st.session_state[pre_keys["removed_fwd"]] = list(
                    saved_sheet_state.get("removed_fwd", [])
                )
                st.session_state[pre_keys["removed_bwd"]] = list(
                    saved_sheet_state.get("removed_bwd", [])
                )
                st.session_state[
                    pre_keys["force_auto_peak_fwd"]
                ] = False
                st.session_state[
                    pre_keys["force_auto_peak_bwd"]
                ] = False

                restore_key_map = {
                    pre_keys["peak_slider_fwd"]:
                        saved_sheet_state.get("peak_vg_fwd"),
                    pre_keys["peak_slider_bwd"]:
                        saved_sheet_state.get("peak_vg_rev"),
                    (
                        f"on_slider_fwd_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): saved_sheet_state.get("on_vg_fwd"),
                    (
                        f"on_slider_rev_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): saved_sheet_state.get("on_vg_rev"),
                    (
                        f"off_slider_fwd_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): saved_sheet_state.get("off_vg_fwd"),
                    (
                        f"off_slider_rev_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): saved_sheet_state.get("off_vg_rev"),
                    pre_keys["ss_slider_fwd"]:
                        saved_sheet_state.get("ss_vg_fwd"),
                    pre_keys["ss_slider_bwd"]:
                        saved_sheet_state.get("ss_vg_rev"),
                    pre_keys["ss_range_start_fwd"]:
                        saved_sheet_state.get("ss_range_start_fwd"),
                    pre_keys["ss_range_end_fwd"]:
                        saved_sheet_state.get("ss_range_end_fwd"),
                    pre_keys["ss_range_start_bwd"]:
                        saved_sheet_state.get("ss_range_start_rev"),
                    pre_keys["ss_range_end_bwd"]:
                        saved_sheet_state.get("ss_range_end_rev"),
                    pre_keys["remove_slider_fwd"]:
                        saved_sheet_state.get(
                            "mobility_remove_vg_fwd"
                        ),
                    pre_keys["remove_slider_bwd"]:
                        saved_sheet_state.get(
                            "mobility_remove_vg_rev"
                        ),
                    pre_keys["log_remove_slider_fwd"]:
                        saved_sheet_state.get("log_remove_vg_fwd"),
                    pre_keys["log_remove_slider_bwd"]:
                        saved_sheet_state.get("log_remove_vg_rev"),
                }

                for restore_key, restore_value in restore_key_map.items():
                    if restore_value is not None:
                        st.session_state[restore_key] = float(
                            restore_value
                        )

                direction_restore_key = (
                    f"direction_view_{file_id}_{selected_sheet}_"
                    f"{operating_mode}"
                )
                saved_direction = saved_sheet_state.get("direction")
                if saved_direction in ("Forward", "Reverse"):
                    st.session_state[
                        direction_restore_key
                    ] = saved_direction

                st.session_state[sheet_restore_token] = True

            elif (
                st.session_state.get("restore_pending")
                and not st.session_state.get(
                    sheet_restore_token, False
                )
            ):
                # Compatibility with older logs that stored only one sheet.
                st.session_state[pre_keys["removed_fwd"]] = list(
                    st.session_state.get("restored_removed_fwd", [])
                )
                st.session_state[pre_keys["removed_bwd"]] = list(
                    st.session_state.get("restored_removed_bwd", [])
                )
                st.session_state[
                    pre_keys["force_auto_peak_fwd"]
                ] = False
                st.session_state[
                    pre_keys["force_auto_peak_bwd"]
                ] = False

                legacy_map = {
                    pre_keys["peak_slider_fwd"]:
                        st.session_state.get("restored_peak_vg_fwd"),
                    pre_keys["peak_slider_bwd"]:
                        st.session_state.get("restored_peak_vg_bwd"),
                    (
                        f"on_slider_fwd_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): st.session_state.get("restored_on_vg_fwd"),
                    (
                        f"on_slider_rev_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): st.session_state.get("restored_on_vg_rev"),
                    (
                        f"off_slider_fwd_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): st.session_state.get("restored_off_vg_fwd"),
                    (
                        f"off_slider_rev_{file_id}_{selected_sheet}_"
                        f"{operating_mode}"
                    ): st.session_state.get("restored_off_vg_rev"),
                    pre_keys["ss_slider_fwd"]:
                        st.session_state.get("restored_ss_vg_fwd"),
                    pre_keys["ss_slider_bwd"]:
                        st.session_state.get("restored_ss_vg_rev"),
                    pre_keys["ss_range_start_fwd"]:
                        st.session_state.get(
                            "restored_ss_range_start_fwd"
                        ),
                    pre_keys["ss_range_end_fwd"]:
                        st.session_state.get(
                            "restored_ss_range_end_fwd"
                        ),
                    pre_keys["ss_range_start_bwd"]:
                        st.session_state.get(
                            "restored_ss_range_start_rev"
                        ),
                    pre_keys["ss_range_end_bwd"]:
                        st.session_state.get(
                            "restored_ss_range_end_rev"
                        ),
                    pre_keys["remove_slider_fwd"]:
                        st.session_state.get(
                            "restored_mobility_remove_vg_fwd"
                        ),
                    pre_keys["remove_slider_bwd"]:
                        st.session_state.get(
                            "restored_mobility_remove_vg_rev"
                        ),
                    pre_keys["log_remove_slider_fwd"]:
                        st.session_state.get(
                            "restored_log_remove_vg_fwd"
                        ),
                    pre_keys["log_remove_slider_bwd"]:
                        st.session_state.get(
                            "restored_log_remove_vg_rev"
                        ),
                }

                for restore_key, restore_value in legacy_map.items():
                    if restore_value is not None:
                        st.session_state[restore_key] = float(
                            restore_value
                        )

                st.session_state[sheet_restore_token] = True

            # Do not clear restored_sheet_states here. Each worksheet restores
            # itself once when the user selects it.

            try:
                res = analyze_sheet(df, file_id, selected_sheet)
            except Exception as exc:
                st.error(str(exc))
                st.stop()

            keys = res["keys"]
            fwd = res["fwd"]
            bwd = res["bwd"]

            # ====================================================
            # ====================================================
            # ====================================================
            # Prepare analysis selections; widgets are rendered next to
            # their corresponding parameter/plot sections below.
            # ====================================================
            initialize_slider_in_range(
                keys["peak_slider_fwd"], fwd,
                float(fwd["GateV"].iloc[res["mobility_max_idx_f"]]),
            )
            initialize_slider_in_range(
                keys["peak_slider_bwd"], bwd,
                float(bwd["GateV"].iloc[res["mobility_max_idx_b"]]),
            )
            initialize_slider_in_range(
                keys["current_slider_fwd"], fwd,
                float(fwd["GateV"].iloc[0]),
            )
            initialize_slider_in_range(
                keys["current_slider_bwd"], bwd,
                float(bwd["GateV"].iloc[0]),
            )

            peak_f_vg = float(st.session_state[keys["peak_slider_fwd"]])
            peak_b_vg = float(st.session_state[keys["peak_slider_bwd"]])
            idx_f = int((fwd["GateV"] - peak_f_vg).abs().idxmin())
            idx_b = int((bwd["GateV"] - peak_b_vg).abs().idxmin())

            params = parameter_values(
                fwd["GateV"], fwd["DrainI_active"],
                res["gm_fwd"], res["mu_fwd"], idx_f,
                bwd["GateV"], bwd["DrainI_active"],
                res["gm_bwd"], res["mu_bwd"], idx_b,
                operating_mode, W,
            )

            current_f_vg = float(st.session_state[keys["current_slider_fwd"]])
            current_b_vg = float(st.session_state[keys["current_slider_bwd"]])
            current_f_idx, current_f_row, current_f_density = current_density_at_vg(
                fwd, current_f_vg, W
            )
            current_b_idx, current_b_row, current_b_density = current_density_at_vg(
                bwd, current_b_vg, W
            )

            # Ensure both Forward and Reverse Peak Elimination selectors
            # exist before reading them. This is required even when only one
            # direction is currently selected in the UI.
            initialize_slider_in_range(
                keys["remove_slider_fwd"],
                fwd,
                float(fwd["GateV"].iloc[res["auto_idx_f"]]),
            )
            initialize_slider_in_range(
                keys["remove_slider_bwd"],
                bwd,
                float(bwd["GateV"].iloc[res["auto_idx_b"]]),
            )
            initialize_slider_in_range(
                keys["log_remove_slider_fwd"],
                fwd,
                float(fwd["GateV"].iloc[res["auto_idx_f"]]),
            )
            initialize_slider_in_range(
                keys["log_remove_slider_bwd"],
                bwd,
                float(bwd["GateV"].iloc[res["auto_idx_b"]]),
            )

            selected_f_idx, selected_f_row = snapped_slider_point(
                fwd, keys["remove_slider_fwd"]
            )
            selected_b_idx, selected_b_row = snapped_slider_point(
                bwd, keys["remove_slider_bwd"]
            )
            selected_f_vg = float(selected_f_row["GateV"])
            selected_b_vg = float(selected_b_row["GateV"])
            selected_f_mu = float(res["mu_fwd"][selected_f_idx])
            selected_b_mu = float(res["mu_bwd"][selected_b_idx])

            removed_f_count = len(st.session_state[keys["removed_fwd"]])
            removed_b_count = len(st.session_state[keys["removed_bwd"]])

            # ====================================================
            # ====================================================
            # ====================================================
            # ====================================================
            # ====================================================
            # Forward and reverse parameters shown simultaneously
            # ====================================================

            def prepare_direction_state(name, short, sweep_df, mu_curve, gm_curve, auto_idx):
                peak_key = (
                    keys["peak_slider_fwd"] if name == "Forward"
                    else keys["peak_slider_bwd"]
                )
                finite_mu = np.where(
                    np.isfinite(mu_curve), mu_curve, -np.inf
                )
                mobility_max_idx = (
                    int(np.argmax(finite_mu))
                    if np.any(np.isfinite(mu_curve)) else 0
                )
                peak_default = float(
                    sweep_df["GateV"].iloc[mobility_max_idx]
                )
                initialize_slider_in_range(
                    peak_key, sweep_df, peak_default
                )

                # Vth is always extracted from the tangent at the mobility-selected Vg.
                vth_key = None

                ss_values = ss_curve(
                    sweep_df["DrainI_active"],
                    sweep_df["GateV"],
                )
                ss_key = (
                    keys["ss_slider_fwd"]
                    if name == "Forward"
                    else keys["ss_slider_bwd"]
                )
                ss_range_start_key = (
                    keys["ss_range_start_fwd"]
                    if name == "Forward"
                    else keys["ss_range_start_bwd"]
                )
                ss_range_end_key = (
                    keys["ss_range_end_fwd"]
                    if name == "Forward"
                    else keys["ss_range_end_bwd"]
                )

                vg_numeric = pd.to_numeric(
                    sweep_df["GateV"], errors="coerce"
                ).to_numpy(dtype=float)
                vg_min = float(np.nanmin(vg_numeric))
                vg_max = float(np.nanmax(vg_numeric))

                if ss_range_start_key not in st.session_state:
                    st.session_state[ss_range_start_key] = vg_min
                if ss_range_end_key not in st.session_state:
                    st.session_state[ss_range_end_key] = vg_max

                raw_ss_start = float(
                    st.session_state.get(ss_range_start_key, vg_min)
                )
                raw_ss_end = float(
                    st.session_state.get(ss_range_end_key, vg_max)
                )
                ss_range_low = max(vg_min, min(raw_ss_start, raw_ss_end))
                ss_range_high = min(vg_max, max(raw_ss_start, raw_ss_end))

                finite_ss_mask = np.isfinite(ss_values)
                range_mask = (
                    finite_ss_mask
                    & np.isfinite(vg_numeric)
                    & (vg_numeric >= ss_range_low)
                    & (vg_numeric <= ss_range_high)
                )
                if np.any(range_mask):
                    candidate_indices = np.where(range_mask)[0]
                    ss_auto_idx = int(
                        candidate_indices[
                            np.argmin(ss_values[candidate_indices])
                        ]
                    )
                elif np.any(finite_ss_mask):
                    candidate_indices = np.where(finite_ss_mask)[0]
                    ss_auto_idx = int(
                        candidate_indices[
                            np.argmin(ss_values[candidate_indices])
                        ]
                    )
                else:
                    ss_auto_idx = auto_idx

                ss_default = float(
                    sweep_df["GateV"].iloc[ss_auto_idx]
                )

                ss_range_signature_key = (
                    f"ss_range_signature_{short}_{file_id}_"
                    f"{selected_sheet}_{operating_mode}"
                )
                ss_range_signature = (
                    round(raw_ss_start, 12),
                    round(raw_ss_end, 12),
                )
                if (
                    ss_key in st.session_state
                    and st.session_state.get(ss_range_signature_key)
                    != ss_range_signature
                ):
                    st.session_state[ss_key] = ss_default
                st.session_state[
                    ss_range_signature_key
                ] = ss_range_signature

                abs_current = np.abs(
                    pd.to_numeric(
                        sweep_df["DrainI_active"], errors="coerce"
                    ).to_numpy(dtype=float)
                )
                on_auto_idx = int(np.nanargmax(abs_current))
                positive = np.isfinite(abs_current) & (abs_current > 0)
                if positive.any():
                    valid_indices = np.where(positive)[0]
                    off_auto_idx = int(
                        valid_indices[np.argmin(abs_current[positive])]
                    )
                else:
                    off_auto_idx = on_auto_idx

                on_key = (
                    f"on_slider_{short}_{file_id}_{selected_sheet}_"
                    f"{operating_mode}"
                )
                off_key = (
                    f"off_slider_{short}_{file_id}_{selected_sheet}_"
                    f"{operating_mode}"
                )
                remove_key = (
                    keys["remove_slider_fwd"]
                    if name == "Forward"
                    else keys["remove_slider_bwd"]
                )
                log_remove_key = (
                    keys["log_remove_slider_fwd"]
                    if name == "Forward"
                    else keys["log_remove_slider_bwd"]
                )

                # Direction switching must never reset either sweep's values.
                # Initialize once, and recalculate only after an actual
                # Peak Elimination/Reset operation changes the active dataset.
                removed_key = (
                    keys["removed_fwd"]
                    if name == "Forward"
                    else keys["removed_bwd"]
                )
                recalc_key = f"recalc_after_change_{removed_key}"
                needs_initialization = any(
                    key not in st.session_state
                    for key in (
                        on_key,
                        off_key,
                        peak_key,
                        remove_key,
                        log_remove_key,
                        ss_key,
                    )
                )
                needs_recalculation = bool(
                    st.session_state.pop(recalc_key, False)
                )

                if needs_initialization or needs_recalculation:
                    st.session_state[on_key] = float(
                        sweep_df["GateV"].iloc[on_auto_idx]
                    )
                    st.session_state[off_key] = float(
                        sweep_df["GateV"].iloc[off_auto_idx]
                    )
                    st.session_state[peak_key] = float(
                        sweep_df["GateV"].iloc[mobility_max_idx]
                    )
                    st.session_state[remove_key] = float(
                        sweep_df["GateV"].iloc[auto_idx]
                    )
                    st.session_state[log_remove_key] = float(
                        sweep_df["GateV"].iloc[auto_idx]
                    )
                    st.session_state[ss_key] = ss_default

                initialize_slider_in_range(
                    on_key,
                    sweep_df,
                    float(sweep_df["GateV"].iloc[on_auto_idx]),
                )
                initialize_slider_in_range(
                    off_key,
                    sweep_df,
                    float(sweep_df["GateV"].iloc[off_auto_idx]),
                )
                initialize_slider_in_range(
                    peak_key,
                    sweep_df,
                    peak_default,
                )
                initialize_slider_in_range(
                    remove_key,
                    sweep_df,
                    float(sweep_df["GateV"].iloc[auto_idx]),
                )
                initialize_slider_in_range(
                    log_remove_key,
                    sweep_df,
                    float(sweep_df["GateV"].iloc[auto_idx]),
                )
                initialize_slider_in_range(
                    ss_key,
                    sweep_df,
                    ss_default,
                )

                peak_vg = float(st.session_state[peak_key])
                peak_idx = int((sweep_df["GateV"] - peak_vg).abs().idxmin())

                vg_for_mu_slope = pd.to_numeric(
                    sweep_df["GateV"], errors="coerce"
                ).to_numpy(dtype=float)
                mu_for_slope = np.asarray(mu_curve, dtype=float)
                with np.errstate(divide="ignore", invalid="ignore"):
                    mobility_slope_abs_curve = np.abs(
                        np.gradient(mu_for_slope, vg_for_mu_slope)
                    )
                mobility_slope_abs_curve[
                    ~np.isfinite(mobility_slope_abs_curve)
                ] = np.nan
                mobility_slope_abs = (
                    float(mobility_slope_abs_curve[peak_idx])
                    if (
                        0 <= peak_idx < len(mobility_slope_abs_curve)
                        and np.isfinite(mobility_slope_abs_curve[peak_idx])
                    )
                    else np.nan
                )

                vth_idx = peak_idx
                vth_vg = float(sweep_df["GateV"].iloc[vth_idx])
                vth_value = vth_at_index(
                    sweep_df["GateV"],
                    sweep_df["DrainI_active"],
                    gm_curve,
                    vth_idx,
                    operating_mode,
                )

                ss_target_vg = float(st.session_state[ss_key])
                ss_idx = int(
                    (sweep_df["GateV"] - ss_target_vg).abs().idxmin()
                )
                st.session_state[ss_key] = float(
                    sweep_df["GateV"].iloc[ss_idx]
                )
                ss_vg = float(sweep_df["GateV"].iloc[ss_idx])
                ss_value = (
                    float(ss_values[ss_idx])
                    if np.isfinite(ss_values[ss_idx]) else np.nan
                )

                on_vg = float(st.session_state[on_key])
                off_vg = float(st.session_state[off_key])
                on_idx, on_row, on_density = current_density_at_vg(
                    sweep_df, on_vg, W
                )
                off_idx, off_row, off_density = current_density_at_vg(
                    sweep_df, off_vg, W
                )
                on_current = abs(float(on_row["DrainI_active"]))
                off_current = abs(float(off_row["DrainI_active"]))
                onoff_value = (
                    on_current / off_current
                    if np.isfinite(off_current) and off_current > 0
                    else np.nan
                )

                return {
                    "name": name,
                    "short": short,
                    "df": sweep_df,
                    "mu_curve": np.asarray(mu_curve, dtype=float),
                    "gm_curve": np.asarray(gm_curve, dtype=float),
                    "ss_curve": np.asarray(ss_values, dtype=float),
                    "auto_idx": auto_idx,
                    "peak_key": peak_key,
                    "peak_default": peak_default,
                    "peak_idx": peak_idx,
                    "peak_vg": float(sweep_df["GateV"].iloc[peak_idx]),
                    "mobility": float(mu_curve[peak_idx]),
                    "mobility_slope_abs": mobility_slope_abs,
                    "vth_key": vth_key,
                    "vth_idx": vth_idx,
                    "vth_vg": float(sweep_df["GateV"].iloc[vth_idx]),
                    "vth": vth_value,
                    "ss_key": ss_key,
                    "ss_default": ss_default,
                    "ss_range_start_key": ss_range_start_key,
                    "ss_range_end_key": ss_range_end_key,
                    "ss_range_low": ss_range_low,
                    "ss_range_high": ss_range_high,
                    "ss_auto_idx": ss_auto_idx,
                    "ss_idx": ss_idx,
                    "ss_vg": float(sweep_df["GateV"].iloc[ss_idx]),
                    "ss": ss_value,
                    "on_key": on_key,
                    "off_key": off_key,
                    "remove_key": remove_key,
                    "log_remove_key": log_remove_key,
                    "on_auto_idx": on_auto_idx,
                    "off_auto_idx": off_auto_idx,
                    "on_idx": on_idx,
                    "off_idx": off_idx,
                    "on_row": on_row,
                    "off_row": off_row,
                    "on_density": on_density,
                    "off_density": off_density,
                    "onoff": onoff_value,
                    "color": "blue" if name == "Forward" else "red",
                }

            f_state = prepare_direction_state(
                "Forward", "fwd", fwd, res["mu_fwd"], res["gm_fwd"],
                res["auto_idx_f"],
            )
            r_state = prepare_direction_state(
                "Reverse", "rev", bwd, res["mu_bwd"], res["gm_bwd"],
                res["auto_idx_b"],
            )

            selected_hysteresis = (
                abs(f_state["vth"] - r_state["vth"])
                if np.isfinite(f_state["vth"]) and np.isfinite(r_state["vth"])
                else np.nan
            )

            # Data title and direction selector.
            header_title_col, header_direction_col = st.columns(
                [5.2, 1.8], gap="small"
            )
            header_title_col.markdown(
                f"<h3 style='color:#333;margin:2px 0 4px 0;'>"
                f"📊 {selected_sheet} ({operating_mode})</h3>",
                unsafe_allow_html=True,
            )
            direction_widget_key = (
                f"direction_view_{file_id}_{selected_sheet}_"
                f"{operating_mode}"
            )
            restored_direction = st.session_state.pop(
                "restored_selected_direction", None
            )
            if restored_direction in ("Forward", "Reverse"):
                st.session_state[direction_widget_key] = restored_direction
            elif direction_widget_key not in st.session_state:
                st.session_state[direction_widget_key] = "Forward"

            selected_direction = header_direction_col.radio(
                "Direction",
                ["Forward", "Reverse"],
                horizontal=True,
                key=direction_widget_key,
                label_visibility="collapsed",
            )
            active_state = (
                f_state if selected_direction == "Forward" else r_state
            )

            # Persistent model-state regression guard. These keys are not tied
            # to visible widgets, so repeated Forward/Reverse switching cannot
            # remove or overwrite the opposite direction's values.
            for guarded_state in (f_state, r_state):
                guarded_df = guarded_state["df"]

                on_default_vg = float(
                    guarded_df["GateV"].iloc[
                        guarded_state["on_auto_idx"]
                    ]
                )
                off_default_vg = float(
                    guarded_df["GateV"].iloc[
                        guarded_state["off_auto_idx"]
                    ]
                )
                elimination_default_vg = float(
                    guarded_df["GateV"].iloc[
                        guarded_state["auto_idx"]
                    ]
                )

                initialize_slider_in_range(
                    guarded_state["on_key"],
                    guarded_df,
                    on_default_vg,
                )
                initialize_slider_in_range(
                    guarded_state["off_key"],
                    guarded_df,
                    off_default_vg,
                )
                initialize_slider_in_range(
                    guarded_state["peak_key"],
                    guarded_df,
                    guarded_state["peak_default"],
                )
                initialize_slider_in_range(
                    guarded_state["ss_key"],
                    guarded_df,
                    guarded_state["ss_default"],
                )
                initialize_slider_in_range(
                    guarded_state["remove_key"],
                    guarded_df,
                    elimination_default_vg,
                )
                initialize_slider_in_range(
                    guarded_state["log_remove_key"],
                    guarded_df,
                    elimination_default_vg,
                )

            def collect_all_sheet_parameters():
                """Calculate every data sheet in original workbook order.

                Visited sheets use their current slider/removal state. Sheets
                that were not opened use the same automatic defaults as the UI.
                Calc and Settings are excluded through target_sheets.
                """
                results = []

                for sheet_order, sheet_name in enumerate(target_sheets):
                    try:
                        sheet_df = standardize_measurement_columns(
                            cached_read_excel_sheet(
                                active_excel_bytes,
                                sheet_name,
                            ).copy()
                        )
                        if not {"GateV", "DrainI"}.issubset(
                            sheet_df.columns
                        ):
                            continue

                        if "DrainV" not in sheet_df.columns:
                            sheet_vd = cached_settings_drain_v(
                                active_excel_bytes,
                                sheet_name,
                            )
                            if sheet_vd is None:
                                continue
                            sheet_df["DrainV"] = float(sheet_vd)

                        sheet_res = analyze_sheet(
                            sheet_df,
                            file_id,
                            sheet_name,
                        )
                        sheet_keys = sheet_res["keys"]

                        def one_direction(
                            direction_name,
                            short_name,
                            sweep_df,
                            mu_curve,
                            gm_curve,
                            auto_idx,
                            mobility_max_idx,
                        ):
                            is_forward = (
                                direction_name == "Forward"
                            )
                            peak_key = (
                                sheet_keys["peak_slider_fwd"]
                                if is_forward
                                else sheet_keys["peak_slider_bwd"]
                            )
                            ss_key = (
                                sheet_keys["ss_slider_fwd"]
                                if is_forward
                                else sheet_keys["ss_slider_bwd"]
                            )
                            ss_start_key = (
                                sheet_keys["ss_range_start_fwd"]
                                if is_forward
                                else sheet_keys["ss_range_start_bwd"]
                            )
                            ss_end_key = (
                                sheet_keys["ss_range_end_fwd"]
                                if is_forward
                                else sheet_keys["ss_range_end_bwd"]
                            )

                            peak_default = float(
                                sweep_df["GateV"].iloc[
                                    mobility_max_idx
                                ]
                            )
                            initialize_slider_in_range(
                                peak_key,
                                sweep_df,
                                peak_default,
                            )
                            peak_idx = int(
                                (
                                    sweep_df["GateV"]
                                    - float(
                                        st.session_state[peak_key]
                                    )
                                ).abs().idxmin()
                            )
                            peak_vg = float(
                                sweep_df["GateV"].iloc[peak_idx]
                            )
                            st.session_state[peak_key] = peak_vg

                            current_abs = np.abs(
                                pd.to_numeric(
                                    sweep_df["DrainI_active"],
                                    errors="coerce",
                                ).to_numpy(dtype=float)
                            )
                            on_auto_idx = int(
                                np.nanargmax(current_abs)
                            )
                            positive = (
                                np.isfinite(current_abs)
                                & (current_abs > 0)
                            )
                            if positive.any():
                                positive_idx = np.where(positive)[0]
                                off_auto_idx = int(
                                    positive_idx[
                                        np.argmin(
                                            current_abs[positive]
                                        )
                                    ]
                                )
                            else:
                                off_auto_idx = on_auto_idx

                            on_key = (
                                f"on_slider_{short_name}_{file_id}_"
                                f"{sheet_name}_{operating_mode}"
                            )
                            off_key = (
                                f"off_slider_{short_name}_{file_id}_"
                                f"{sheet_name}_{operating_mode}"
                            )
                            initialize_slider_in_range(
                                on_key,
                                sweep_df,
                                float(
                                    sweep_df["GateV"].iloc[
                                        on_auto_idx
                                    ]
                                ),
                            )
                            initialize_slider_in_range(
                                off_key,
                                sweep_df,
                                float(
                                    sweep_df["GateV"].iloc[
                                        off_auto_idx
                                    ]
                                ),
                            )
                            _, on_row, on_density = (
                                current_density_at_vg(
                                    sweep_df,
                                    st.session_state[on_key],
                                    W,
                                )
                            )
                            _, off_row, off_density = (
                                current_density_at_vg(
                                    sweep_df,
                                    st.session_state[off_key],
                                    W,
                                )
                            )
                            on_current = abs(
                                float(on_row["DrainI_active"])
                            )
                            off_current = abs(
                                float(off_row["DrainI_active"])
                            )
                            onoff = (
                                on_current / off_current
                                if np.isfinite(off_current)
                                and off_current > 0
                                else np.nan
                            )

                            ss_values = ss_curve(
                                sweep_df["DrainI_active"],
                                sweep_df["GateV"],
                            )
                            vg_values = pd.to_numeric(
                                sweep_df["GateV"],
                                errors="coerce",
                            ).to_numpy(dtype=float)
                            vg_min = float(np.nanmin(vg_values))
                            vg_max = float(np.nanmax(vg_values))
                            st.session_state.setdefault(
                                ss_start_key,
                                vg_min,
                            )
                            st.session_state.setdefault(
                                ss_end_key,
                                vg_max,
                            )
                            ss_start = float(
                                st.session_state[ss_start_key]
                            )
                            ss_end = float(
                                st.session_state[ss_end_key]
                            )
                            low = max(
                                vg_min,
                                min(ss_start, ss_end),
                            )
                            high = min(
                                vg_max,
                                max(ss_start, ss_end),
                            )
                            valid_ss = (
                                np.isfinite(ss_values)
                                & (vg_values >= low)
                                & (vg_values <= high)
                            )
                            if valid_ss.any():
                                candidates = np.where(valid_ss)[0]
                                ss_auto_idx = int(
                                    candidates[
                                        np.argmin(
                                            ss_values[candidates]
                                        )
                                    ]
                                )
                            else:
                                finite_ss = np.where(
                                    np.isfinite(ss_values)
                                )[0]
                                ss_auto_idx = (
                                    int(finite_ss[0])
                                    if len(finite_ss)
                                    else auto_idx
                                )
                            ss_default = float(
                                sweep_df["GateV"].iloc[
                                    ss_auto_idx
                                ]
                            )
                            initialize_slider_in_range(
                                ss_key,
                                sweep_df,
                                ss_default,
                            )
                            ss_idx = int(
                                (
                                    sweep_df["GateV"]
                                    - float(
                                        st.session_state[ss_key]
                                    )
                                ).abs().idxmin()
                            )
                            ss_value = (
                                float(ss_values[ss_idx])
                                if np.isfinite(ss_values[ss_idx])
                                else np.nan
                            )

                            vth_value = vth_at_index(
                                sweep_df["GateV"],
                                sweep_df["DrainI_active"],
                                gm_curve,
                                peak_idx,
                                operating_mode,
                            )

                            return {
                                "mobility": float(
                                    mu_curve[peak_idx]
                                ),
                                "vth": float(vth_value),
                                "ss": float(ss_value),
                                "on_density": float(on_density),
                                "off_density": float(off_density),
                                "onoff": float(onoff),
                            }

                        f_values = one_direction(
                            "Forward",
                            "fwd",
                            sheet_res["fwd"],
                            sheet_res["mu_fwd"],
                            sheet_res["gm_fwd"],
                            sheet_res["auto_idx_f"],
                            sheet_res["mobility_max_idx_f"],
                        )
                        r_values = one_direction(
                            "Reverse",
                            "rev",
                            sheet_res["bwd"],
                            sheet_res["mu_bwd"],
                            sheet_res["gm_bwd"],
                            sheet_res["auto_idx_b"],
                            sheet_res["mobility_max_idx_b"],
                        )

                        fwd_df = sheet_res["fwd"]
                        bwd_df = sheet_res["bwd"]
                        results.append({
                            "Sheet": sheet_name,
                            "_sheet_order": sheet_order,
                            "Operating mode": operating_mode,
                            "Drain voltage (V)": float(
                                sheet_res["vd"]
                            ),
                            "Gate voltage range": (
                                f"{float(min(fwd_df['GateV'].min(), bwd_df['GateV'].min())):.2f} "
                                f"to {float(max(fwd_df['GateV'].max(), bwd_df['GateV'].max())):.2f} V"
                            ),
                            "Gate voltage step (V)": (
                                float(
                                    np.median(
                                        np.abs(
                                            np.diff(
                                                fwd_df["GateV"]
                                            )
                                        )
                                    )
                                )
                                if len(fwd_df) > 1
                                else np.nan
                            ),
                            "Forward mobility (cm²/V·s)": (
                                f_values["mobility"]
                            ),
                            "Forward Vth (V)": f_values["vth"],
                            "Forward SS (mV/dec)": f_values["ss"],
                            "Forward ON/OFF ratio": (
                                f_values["onoff"]
                            ),
                            "Forward ON current / Width (A/μm)": (
                                f_values["on_density"]
                            ),
                            "Forward OFF current / Width (A/μm)": (
                                f_values["off_density"]
                            ),
                            "Backward mobility (cm²/V·s)": (
                                r_values["mobility"]
                            ),
                            "Backward Vth (V)": r_values["vth"],
                            "Backward SS (mV/dec)": r_values["ss"],
                            "Backward ON/OFF ratio": (
                                r_values["onoff"]
                            ),
                            "Backward ON current / Width (A/μm)": (
                                r_values["on_density"]
                            ),
                            "Backward OFF current / Width (A/μm)": (
                                r_values["off_density"]
                            ),
                            "Hysteresis (V)": (
                                abs(
                                    f_values["vth"]
                                    - r_values["vth"]
                                )
                                if np.isfinite(f_values["vth"])
                                and np.isfinite(r_values["vth"])
                                else np.nan
                            ),
                            "ON/OFF ratio": f_values["onoff"],
                            "ON current / Width (A/μm)": (
                                f_values["on_density"]
                            ),
                            "OFF current / Width (A/μm)": (
                                f_values["off_density"]
                            ),
                        })
                    except Exception:
                        # Invalid/non-measurement sheets are skipped.
                        continue

                return results


            def collect_all_sheet_states():
                """Capture all worksheet-specific editing states.

                Session-state keys already include file, sheet and mode, so
                each worksheet can be saved independently in one project log.
                """
                all_states = {}

                for sheet_name in target_sheets:
                    sheet_keys = state_keys(
                        file_id,
                        sheet_name,
                        operating_mode,
                    )

                    def get_number(key):
                        value = st.session_state.get(key)
                        if value is None:
                            return None
                        try:
                            return float(value)
                        except (TypeError, ValueError):
                            return None

                    direction_key = (
                        f"direction_view_{file_id}_{sheet_name}_"
                        f"{operating_mode}"
                    )
                    on_fwd_key = (
                        f"on_slider_fwd_{file_id}_{sheet_name}_"
                        f"{operating_mode}"
                    )
                    on_rev_key = (
                        f"on_slider_rev_{file_id}_{sheet_name}_"
                        f"{operating_mode}"
                    )
                    off_fwd_key = (
                        f"off_slider_fwd_{file_id}_{sheet_name}_"
                        f"{operating_mode}"
                    )
                    off_rev_key = (
                        f"off_slider_rev_{file_id}_{sheet_name}_"
                        f"{operating_mode}"
                    )

                    all_states[sheet_name] = {
                        "removed_fwd": list(
                            st.session_state.get(
                                sheet_keys["removed_fwd"], []
                            )
                        ),
                        "removed_bwd": list(
                            st.session_state.get(
                                sheet_keys["removed_bwd"], []
                            )
                        ),
                        "peak_vg_fwd": get_number(
                            sheet_keys["peak_slider_fwd"]
                        ),
                        "peak_vg_rev": get_number(
                            sheet_keys["peak_slider_bwd"]
                        ),
                        "on_vg_fwd": get_number(on_fwd_key),
                        "on_vg_rev": get_number(on_rev_key),
                        "off_vg_fwd": get_number(off_fwd_key),
                        "off_vg_rev": get_number(off_rev_key),
                        "ss_vg_fwd": get_number(
                            sheet_keys["ss_slider_fwd"]
                        ),
                        "ss_vg_rev": get_number(
                            sheet_keys["ss_slider_bwd"]
                        ),
                        "ss_range_start_fwd": get_number(
                            sheet_keys["ss_range_start_fwd"]
                        ),
                        "ss_range_end_fwd": get_number(
                            sheet_keys["ss_range_end_fwd"]
                        ),
                        "ss_range_start_rev": get_number(
                            sheet_keys["ss_range_start_bwd"]
                        ),
                        "ss_range_end_rev": get_number(
                            sheet_keys["ss_range_end_bwd"]
                        ),
                        "mobility_remove_vg_fwd": get_number(
                            sheet_keys["remove_slider_fwd"]
                        ),
                        "mobility_remove_vg_rev": get_number(
                            sheet_keys["remove_slider_bwd"]
                        ),
                        "log_remove_vg_fwd": get_number(
                            sheet_keys["log_remove_slider_fwd"]
                        ),
                        "log_remove_vg_rev": get_number(
                            sheet_keys["log_remove_slider_bwd"]
                        ),
                        "direction": st.session_state.get(
                            direction_key, "Forward"
                        ),
                    }

                return all_states


            def collect_all_plot_snapshots():
                """Return saved Plotly JSON for every visited worksheet.

                A worksheet snapshot contains curves, markers, tangent lines,
                selected direction styling, axis ranges and annotations.
                Unvisited worksheets are reconstructed from their saved
                analysis state when opened.
                """
                snapshots = {}
                for sheet_name in target_sheets:
                    snapshot_key = (
                        f"plot_snapshot_{file_id}_{sheet_name}_"
                        f"{operating_mode}"
                    )
                    snapshot = st.session_state.get(snapshot_key)
                    if snapshot:
                        snapshots[sheet_name] = snapshot

                # Preserve snapshots from a previously opened log even when a
                # sheet was not revisited during the current editing session.
                restored = st.session_state.get(
                    "restored_plot_snapshots", {}
                )
                if isinstance(restored, dict):
                    for sheet_name, snapshot in restored.items():
                        snapshots.setdefault(sheet_name, snapshot)

                return snapshots


            def build_current_log_entry(log_id=None):
                if current_project:
                    current_settings = ensure_project_device_settings(
                        current_project
                    )
                    current_settings["operating_mode"] = operating_mode
                    current_settings["width_um"] = float(W)
                    current_settings["length_um"] = float(L)
                    current_settings["cox_nf_cm2"] = float(Cox_nf)
                try:
                    uploaded_file.seek(0)
                    saved_file_bytes = uploaded_file.read()
                    uploaded_file.seek(0)
                except Exception:
                    saved_file_bytes = st.session_state.get(
                        "active_file_bytes"
                    )

                active_analysis_file_name = getattr(
                    uploaded_file,
                    "name",
                    st.session_state.get(
                        "active_file_name", "restored.xlsx"
                    ),
                )

                all_sheet_parameters = collect_all_sheet_parameters()
                all_sheet_states = collect_all_sheet_states()
                all_plot_snapshots = collect_all_plot_snapshots()

                return {
                    "_log_id": log_id or str(uuid.uuid4()),
                    "_project_name": current_project,
                    "_selected_direction": selected_direction,
                    "_file_bytes": saved_file_bytes,
                    "_removed_fwd_indices": list(
                        st.session_state[keys["removed_fwd"]]
                    ),
                    "_removed_bwd_indices": list(
                        st.session_state[keys["removed_bwd"]]
                    ),
                    "_on_vg_fwd": float(f_state["on_row"]["GateV"]),
                    "_on_vg_rev": float(r_state["on_row"]["GateV"]),
                    "_off_vg_fwd": float(f_state["off_row"]["GateV"]),
                    "_off_vg_rev": float(r_state["off_row"]["GateV"]),
                    "_vth_vg_fwd": float(f_state["vth_vg"]),
                    "_vth_vg_rev": float(r_state["vth_vg"]),
                    "_ss_vg_fwd": float(f_state["ss_vg"]),
                    "_ss_vg_rev": float(r_state["ss_vg"]),
                    "_ss_range_start_fwd": float(
                        st.session_state[f_state["ss_range_start_key"]]
                    ),
                    "_ss_range_end_fwd": float(
                        st.session_state[f_state["ss_range_end_key"]]
                    ),
                    "_ss_range_start_rev": float(
                        st.session_state[r_state["ss_range_start_key"]]
                    ),
                    "_ss_range_end_rev": float(
                        st.session_state[r_state["ss_range_end_key"]]
                    ),
                    "_mobility_remove_vg_fwd": float(
                        st.session_state[f_state["remove_key"]]
                    ),
                    "_mobility_remove_vg_rev": float(
                        st.session_state[r_state["remove_key"]]
                    ),
                    "_log_remove_vg_fwd": float(
                        st.session_state[f_state["log_remove_key"]]
                    ),
                    "_log_remove_vg_rev": float(
                        st.session_state[r_state["log_remove_key"]]
                    ),
                    "Saved at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "File": active_analysis_file_name,
                    "_all_sheet_parameters": all_sheet_parameters,
                    "_sheet_states": all_sheet_states,
                    "_plot_snapshots": all_plot_snapshots,
                    "Sheet": selected_sheet,
                    "Operating mode": operating_mode,
                    "Width (μm)": float(W),
                    "Length (μm)": float(L),
                    "Cox (nF/cm²)": float(Cox_nf),
                    "Drain voltage (V)": float(res["vd"]),
                    "Gate voltage range": (
                        f"{float(min(fwd['GateV'].min(), bwd['GateV'].min())):.2f} "
                        f"to {float(max(fwd['GateV'].max(), bwd['GateV'].max())):.2f} V"
                    ),
                    "Gate voltage step (V)": float(
                        np.median(np.abs(np.diff(fwd["GateV"])))
                    ) if len(fwd) > 1 else np.nan,
                    "Forward mobility (cm²/V·s)": float(f_state["mobility"]),
                    "Forward Vth (V)": float(f_state["vth"]),
                    "Forward peak Vg (V)": float(f_state["peak_vg"]),
                    "Forward SS (mV/dec)": float(f_state["ss"]),
                    "Forward ON/OFF ratio": float(f_state["onoff"]),
                    "Forward ON current / Width (A/μm)": float(f_state["on_density"]),
                    "Forward OFF current / Width (A/μm)": float(f_state["off_density"]),
                    "Backward mobility (cm²/V·s)": float(r_state["mobility"]),
                    "Backward Vth (V)": float(r_state["vth"]),
                    "Backward peak Vg (V)": float(r_state["peak_vg"]),
                    "Backward SS (mV/dec)": float(r_state["ss"]),
                    "Backward ON/OFF ratio": float(r_state["onoff"]),
                    "Backward ON current / Width (A/μm)": float(r_state["on_density"]),
                    "Backward OFF current / Width (A/μm)": float(r_state["off_density"]),
                    "Hysteresis (V)": float(selected_hysteresis),
                    "ON/OFF ratio": float(f_state["onoff"]),
                    "ON current / Width (A/μm)": float(f_state["on_density"]),
                    "OFF current / Width (A/μm)": float(f_state["off_density"]),
                    "Selected Forward Vg (V)": float(f_state["on_row"]["GateV"]),
                    "Selected Backward Vg (V)": float(r_state["on_row"]["GateV"]),
                }

            if st.session_state.pop("save_current_requested", False):
                active_id = st.session_state.get(
                    "persistent_active_log_id"
                )
                target_folder = None
                target_index = None

                # First try the currently selected project.
                current_records = st.session_state[
                    "analysis_log_folders"
                ].get(current_project, [])
                for index, record in enumerate(current_records):
                    if record.get("_log_id") == active_id:
                        target_folder = current_project
                        target_index = index
                        break

                # If the project selector changed or was rebuilt during restore,
                # locate the active log across all projects.
                if target_index is None:
                    for folder_name, records in st.session_state[
                        "analysis_log_folders"
                    ].items():
                        for index, record in enumerate(records):
                            if record.get("_log_id") == active_id:
                                target_folder = folder_name
                                target_index = index
                                break
                        if target_index is not None:
                            break

                if target_folder is not None and target_index is not None:
                    updated_entry = build_current_log_entry(active_id)
                    st.session_state["analysis_log_folders"][
                        target_folder
                    ][target_index] = updated_entry
                    st.session_state["active_log_folder"] = target_folder
                    st.session_state[
                        "persistent_active_log_id"
                    ] = active_id
                    st.session_state[
                        "active_file_bytes"
                    ] = updated_entry["_file_bytes"]
                    st.session_state[
                        "active_file_name"
                    ] = updated_entry.get(
                        "File", "restored.xlsx"
                    )
                    st.session_state["active_file_source"] = "log"
                    update_active_project_workspace(
                        persist=True,
                        active_file_bytes=updated_entry["_file_bytes"],
                        active_file_name=updated_entry.get(
                            "File", "restored.xlsx"
                        ),
                        active_file_source="log",
                        active_log_id=active_id,
                        uploader_generation=int(
                            st.session_state.get(
                                "file_uploader_generation", 0
                            )
                        ),
                        selected_sheet=selected_sheet,
                    )
                    save_projects_state()
                    sheet_count = len(
                        updated_entry.get("_all_sheet_parameters", [])
                    )
                    st.session_state["save_status_message"] = (
                        f"파일 전체 {sheet_count}개 시트의 값을 "
                        "한 번에 저장했습니다."
                    )
                else:
                    st.warning(
                        "활성 로그를 찾지 못했습니다. 로그를 다시 연 뒤 "
                        "Save를 눌러주세요."
                    )

            if st.session_state.get("save_status_message"):
                st.success(st.session_state.pop("save_status_message"))

            if st.session_state.pop("add_project_requested", False):
                log_entry = build_current_log_entry()
                st.session_state["analysis_log_folders"][current_project].append(
                    log_entry
                )
                st.session_state["active_log_folder"] = current_project
                st.session_state["persistent_active_log_id"] = log_entry["_log_id"]
                st.session_state["active_file_bytes"] = log_entry["_file_bytes"]
                st.session_state["active_file_name"] = log_entry.get(
                    "File", "restored.xlsx"
                )
                st.session_state["active_file_source"] = "log"
                st.session_state["file_uploader_generation"] = (
                    int(st.session_state.get("file_uploader_generation", 0)) + 1
                )
                update_active_project_workspace(
                    persist=True,
                    active_file_bytes=log_entry["_file_bytes"],
                    active_file_name=log_entry.get(
                        "File", "restored.xlsx"
                    ),
                    active_file_source="log",
                    active_log_id=log_entry["_log_id"],
                    uploader_generation=int(
                        st.session_state["file_uploader_generation"]
                    ),
                    selected_sheet=selected_sheet,
                )
                save_projects_state()
                sheet_count = len(
                    log_entry.get("_all_sheet_parameters", [])
                )
                st.success(
                    f"'{current_project}' 프로젝트에 파일 1개와 "
                    f"{sheet_count}개 시트 결과를 한 번에 추가했습니다."
                )
                st.rerun()

            def slider_with_auto(container, state, key_name, default_value, label, prefix):
                render_discrete_vg_control(
                    title="",
                    slider_label=label,
                    state_key=key_name,
                    active_df=state["df"],
                    default_value=float(default_value),
                    button_prefix=prefix,
                    parent=container,
                )
                container.button(
                    "Auto Set",
                    key=f"{prefix}_auto_set",
                    use_container_width=True,
                    on_click=set_state_value,
                    args=(key_name, float(default_value)),
                )

            def dual_metric_box(title, f_value, r_value, renderer_f=None, renderer_r=None):
                with st.container(border=False):
                    st.markdown(
                        f"<div class='metric-title'>{title}</div>",
                        unsafe_allow_html=True,
                    )
                    left, right = st.columns([1, 1], gap="medium")
                    with left:
                        st.markdown(
                            f"<div style='color:#2E60AB;font-size:12px;font-weight:750;'>"
                            f"Forward</div><div class='metric-value'>{f_value}</div>",
                            unsafe_allow_html=True,
                        )
                        if renderer_f is not None:
                            renderer_f(left)
                    with right:
                        st.markdown(
                            f"<div style='border-left:1px solid rgba(120,120,120,.28);"
                            f"padding-left:8px;color:#D94B45;font-size:12px;font-weight:750;'>"
                            f"Reverse</div><div class='metric-value' style='border-left:1px solid "
                            f"rgba(120,120,120,.28);padding-left:8px;'>{r_value}</div>",
                            unsafe_allow_html=True,
                        )
                        if renderer_r is not None:
                            renderer_r(right)


            # ====================================================
            # ====================================================
            # ====================================================
            # Direction-aware parameter summary above plots
            # ====================================================
            direction_color = (
                "#2E60AB" if selected_direction == "Forward" else "#D94B45"
            )

            def render_top_parameter(title, value):
                st.markdown(
                    f"<div class='top-param-card'>"
                    f"<div class='top-param-title'>{title}</div>"
                    f"<div class='top-param-value' style='color:{direction_color};'>"
                    f"{value}</div></div>",
                    unsafe_allow_html=True,
                )

            top_row_1 = st.columns(4, gap="large")
            with top_row_1[0]:
                render_top_parameter(
                    "ON Current / Width",
                    f"{sci(active_state['on_density'])} A/μm",
                )
            with top_row_1[1]:
                render_top_parameter(
                    "OFF Current / Width",
                    f"{sci(active_state['off_density'])} A/μm",
                )
            with top_row_1[2]:
                render_top_parameter(
                    "ON/OFF Ratio",
                    sci(active_state["onoff"]),
                )
            with top_row_1[3]:
                render_top_parameter(
                    "Mobility",
                    f"{active_state['mobility']:.2f} cm²/V·s",
                )

            st.markdown(
                "<div class='param-row-gap'></div>",
                unsafe_allow_html=True,
            )

            top_row_2 = st.columns(4, gap="large")
            with top_row_2[0]:
                render_top_parameter(
                    "SS Value",
                    (
                        f"{active_state['ss']:.1f} mV/dec"
                        if np.isfinite(active_state["ss"]) else "N/A"
                    ),
                )
            with top_row_2[1]:
                render_top_parameter(
                    "Threshold Voltage",
                    (
                        f"{active_state['vth']:.2f} V"
                        if np.isfinite(active_state["vth"]) else "N/A"
                    ),
                )
            with top_row_2[2]:
                render_top_parameter(
                    "Hysteresis",
                    (
                        f"{selected_hysteresis:.2f} V"
                        if np.isfinite(selected_hysteresis) else "N/A"
                    ),
                )
            with top_row_2[3]:
                active_slope_value = active_state.get(
                    "mobility_slope_abs", np.nan
                )
                render_top_parameter(
                    "|dμ/dV<sub>G</sub>|",
                    (
                        f"{active_slope_value:.2e} cm² V⁻² s⁻¹"
                        if np.isfinite(active_slope_value)
                        else "N/A"
                    ),
                )

            # ====================================================
            # Four horizontal plots
            # ====================================================
            graph_mobility_title = (
                "Linear Mobility" if operating_mode == "Linear"
                else "Saturation Mobility"
            )
            fig = make_subplots(
                rows=1,
                cols=4,
                subplot_titles=(
                    "Transfer (Log)",
                    "Subthreshold Swing",
                    graph_mobility_title,
                    "Transfer (Linear)",
                ),
                horizontal_spacing=0.085,
            )

            vg_fwd = fwd["GateV"]
            id_fwd = fwd["DrainI_active"]
            vg_bwd = bwd["GateV"]
            id_bwd = bwd["DrainI_active"]

            fwd_dash = (
                "solid" if selected_direction == "Forward" else "dash"
            )
            rev_dash = (
                "solid" if selected_direction == "Reverse" else "dash"
            )

            for col_num in (1, 4):
                fig.add_trace(
                    go.Scatter(
                        x=vg_fwd,
                        y=np.abs(id_fwd),
                        line=dict(color="blue", dash=fwd_dash, width=2),
                        showlegend=False,
                    ),
                    row=1, col=col_num,
                )
                fig.add_trace(
                    go.Scatter(
                        x=vg_bwd,
                        y=np.abs(id_bwd),
                        line=dict(color="red", dash=rev_dash, width=2),
                        showlegend=False,
                    ),
                    row=1, col=col_num,
                )

            if "GateI" in df.columns:
                gate_i = pd.to_numeric(df["GateI"], errors="coerce")
                ig_f = gate_i.iloc[
                    fwd["__source_index"].astype(int).to_numpy()
                ].reset_index(drop=True)
                ig_b = gate_i.iloc[
                    bwd["__source_index"].astype(int).to_numpy()
                ].reset_index(drop=True)
                for col_num in (1, 4):
                    fig.add_trace(
                        go.Scatter(
                            x=vg_fwd,
                            y=np.abs(ig_f),
                            line=dict(
                                color="dimgray",
                                dash=("dot" if selected_direction == "Forward" else "dashdot"),
                            ),
                            showlegend=False,
                        ),
                        row=1, col=col_num,
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=vg_bwd,
                            y=np.abs(ig_b),
                            line=dict(
                                color="black",
                                dash=("dot" if selected_direction == "Reverse" else "dashdot"),
                            ),
                            showlegend=False,
                        ),
                        row=1, col=col_num,
                    )

            # ON/OFF locations on Transfer (Log): color by sweep direction.
            # Forward = blue, Reverse = red for both ON and OFF.
            for marker_state in (f_state, r_state):
                active_width = (
                    1.8
                    if marker_state["name"] == selected_direction
                    else 1.2
                )
                direction_line_color = (
                    "blue"
                    if marker_state["name"] == "Forward"
                    else "red"
                )
                fig.add_vline(
                    x=float(marker_state["on_row"]["GateV"]),
                    line_dash="dot",
                    line_width=active_width,
                    line_color=direction_line_color,
                    row=1,
                    col=1,
                )
                fig.add_vline(
                    x=float(marker_state["off_row"]["GateV"]),
                    line_dash="dot",
                    line_width=active_width,
                    line_color=direction_line_color,
                    row=1,
                    col=1,
                )

            # Transfer-log Peak Elimination targets for both directions.
            log_target_specs = (
                (f_state, keys["log_remove_slider_fwd"]),
                (r_state, keys["log_remove_slider_bwd"]),
            )
            for log_state, log_key in log_target_specs:
                log_idx, log_row = snapped_slider_point(
                    log_state["df"], log_key
                )
                log_x = float(log_row["GateV"])
                log_y = abs(float(log_row["DrainI_active"]))
                fig.add_trace(
                    go.Scatter(
                        x=[log_x],
                        y=[log_y],
                        mode="markers",
                        marker=dict(
                            symbol="x",
                            size=12,
                            color=log_state["color"],
                            line=dict(
                                width=2,
                                color=log_state["color"],
                            ),
                        ),
                        showlegend=False,
                        hovertemplate=(
                            "Vg=%{x:.3f} V"
                            "<br>|Id|=%{y:.3e} A"
                            "<extra></extra>"
                        ),
                    ),
                    row=1,
                    col=1,
                )

            # Mobility curves and independently selected peak lines.
            fig.add_trace(
                go.Scatter(
                    x=vg_fwd,
                    y=res["mu_fwd"],
                    line=dict(color="blue", dash=fwd_dash, width=2),
                    showlegend=False,
                ),
                row=1, col=3,
            )
            fig.add_trace(
                go.Scatter(
                    x=vg_bwd,
                    y=res["mu_bwd"],
                    line=dict(color="red", dash=rev_dash, width=2),
                    showlegend=False,
                ),
                row=1, col=3,
            )
            for marker_state in (f_state, r_state):
                fig.add_vline(
                    x=marker_state["peak_vg"],
                    line_dash="dot",
                    line_width=1.5,
                    line_color=marker_state["color"],
                    row=1, col=3,
                )

            # Peak-elimination targets for both directions.
            for elimination_row, elimination_mu, elimination_color in (
                (selected_f_row, selected_f_mu, "blue"),
                (selected_b_row, selected_b_mu, "red"),
            ):
                fig.add_trace(
                    go.Scatter(
                        x=[float(elimination_row["GateV"])],
                        y=[float(elimination_mu)],
                        mode="markers",
                        marker=dict(
                            symbol="x",
                            size=12,
                            color=elimination_color,
                            line=dict(
                                width=2,
                                color=elimination_color,
                            ),
                        ),
                        showlegend=False,
                        hovertemplate=(
                            "Vg=%{x:.3f} V"
                            "<br>Mobility=%{y:.3e}"
                            "<extra></extra>"
                        ),
                    ),
                    row=1,
                    col=3,
                )

            # SS curves for both directions.
            fig.add_trace(
                go.Scatter(
                    x=f_state["df"]["GateV"],
                    y=f_state["ss_curve"],
                    line=dict(color="blue", dash=fwd_dash, width=2),
                    showlegend=False,
                ),
                row=1, col=2,
            )
            fig.add_trace(
                go.Scatter(
                    x=r_state["df"]["GateV"],
                    y=r_state["ss_curve"],
                    line=dict(color="red", dash=rev_dash, width=2),
                    showlegend=False,
                ),
                row=1, col=2,
            )
            for ss_state in (f_state, r_state):
                if np.isfinite(ss_state["ss"]):
                    fig.add_vline(
                        x=ss_state["ss_vg"],
                        line_dash="dot",
                        line_width=(
                            1.8
                            if ss_state["name"] == selected_direction
                            else 1.2
                        ),
                        line_color=ss_state["color"],
                        row=1,
                        col=2,
                    )

            # Transfer (Linear): restore tangent and selected point.
            for tangent_state in (f_state, r_state):
                x_all = np.asarray(
                    tangent_state["df"]["GateV"],
                    dtype=float,
                )
                y_all = np.abs(
                    np.asarray(
                        tangent_state["df"]["DrainI_active"],
                        dtype=float,
                    )
                )
                tangent_idx = tangent_state["vth_idx"]
                slope_abs = np.gradient(y_all, x_all)[tangent_idx]
                tangent_y = (
                    y_all[tangent_idx]
                    + slope_abs * (x_all - x_all[tangent_idx])
                )
                valid_tangent = (
                    np.isfinite(tangent_y) & (tangent_y >= 0)
                )
                if valid_tangent.sum() >= 2:
                    fig.add_trace(
                        go.Scatter(
                            x=x_all[valid_tangent],
                            y=tangent_y[valid_tangent],
                            line=dict(
                                color=tangent_state["color"],
                                dash=(
                                    "dot"
                                    if tangent_state["name"]
                                    == selected_direction
                                    else "dash"
                                ),
                                width=(
                                    1.8
                                    if tangent_state["name"]
                                    == selected_direction
                                    else 1.2
                                ),
                            ),
                            showlegend=False,
                        ),
                        row=1,
                        col=4,
                    )
                fig.add_trace(
                    go.Scatter(
                        x=[tangent_state["vth_vg"]],
                        y=[
                            abs(
                                float(
                                    tangent_state["df"][
                                        "DrainI_active"
                                    ].iloc[tangent_idx]
                                )
                            )
                        ],
                        mode="markers",
                        marker=dict(
                            size=9,
                            color=tangent_state["color"],
                            symbol="circle",
                            line=dict(width=1, color="white"),
                        ),
                        showlegend=False,
                    ),
                    row=1,
                    col=4,
                )

            common_axis = dict(
                ticks="outside", showline=True, mirror=True, showgrid=True,
                gridcolor="lightgray", griddash="dot", zeroline=False,
            )
            fig.update_xaxes(title_text="Gate Voltage (V)", **common_axis)
            fig.update_yaxes(
                title_text="Current (A)",
                type="log",
                tickformat=".1e",
                exponentformat="E",
                showexponent="all",
                row=1,
                col=1,
                **common_axis,
            )
            fig.update_yaxes(
                title_text="SS (mV/dec)",
                range=[0, 500000],
                tickformat=".1e",
                exponentformat="E",
                showexponent="all",
                row=1,
                col=2,
                **common_axis,
            )
            fig.update_yaxes(
                title_text="Mobility (cm²/V·s)",
                row=1,
                col=3,
                **common_axis,
            )
            fig.update_yaxes(
                title_text="Current (A)",
                tickformat=".1e",
                exponentformat="E",
                showexponent="all",
                row=1,
                col=4,
                **common_axis,
            )
            fig.update_layout(
                height=330,
                template="plotly_white",
                margin=dict(t=34, b=8, l=32, r=8),
                showlegend=False,
            )
            # Save the exact interactive Plotly figure for this worksheet.
            plot_snapshot_key = (
                f"plot_snapshot_{file_id}_{selected_sheet}_"
                f"{operating_mode}"
            )
            current_plot_json = fig.to_json()
            st.session_state[plot_snapshot_key] = current_plot_json

            # On the first render after opening a saved project log, display
            # the saved figure snapshot exactly as it looked at Save time.
            restored_plot_snapshots = st.session_state.get(
                "restored_plot_snapshots", {}
            )
            plot_restore_token = (
                f"plot_snapshot_applied_"
                f"{st.session_state.get('restored_state_log_id', 'none')}_"
                f"{file_id}_{selected_sheet}_{operating_mode}"
            )
            saved_plot_json = (
                restored_plot_snapshots.get(selected_sheet)
                if isinstance(restored_plot_snapshots, dict)
                else None
            )

            figure_to_display = fig
            if (
                saved_plot_json
                and not st.session_state.get(
                    plot_restore_token, False
                )
            ):
                try:
                    figure_to_display = go.Figure(
                        json.loads(saved_plot_json)
                    )
                    st.session_state[plot_restore_token] = True
                except Exception:
                    figure_to_display = fig

            st.plotly_chart(
                figure_to_display,
                use_container_width=True,
            )

            def render_removal_control(state, container, removed_key, remove_key, force_key):
                with container:
                    st.markdown(
                        f"<div style='font-weight:750; color:{state['color']}; "
                        f"font-size:12px; margin-bottom:3px;'>{state['name']}</div>",
                        unsafe_allow_html=True,
                    )
                    default_vg = float(
                        state["df"]["GateV"].iloc[state["auto_idx"]]
                    )
                    initialize_slider_in_range(remove_key, state["df"], default_vg)
                    removal_vg = render_discrete_vg_control(
                        title="",
                        slider_label=f"{state['name']} removal Vg",
                        state_key=remove_key,
                        active_df=state["df"],
                        default_value=default_vg,
                        button_prefix=(
                            f"remove_inline_{state['short']}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        parent=container,
                    )
                    _, removal_row = nearest_row_by_vg(state["df"], removal_vg)
                    remove_col, reset_col = container.columns([1, 1], gap="small")
                    remove_col.button(
                        "Rm",
                        key=(
                            f"remove_inline_btn_{state['short']}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        use_container_width=True,
                        on_click=remove_mobility_point,
                        args=(
                            removed_key,
                            int(removal_row["__source_index"]),
                            force_key,
                            (
                                state["peak_key"], state["on_key"],
                                state["off_key"], state["vth_key"],
                                state["ss_key"], remove_key,
                            ),
                        ),
                    )
                    reset_col.button(
                        "Rst",
                        key=(
                            f"reset_inline_btn_{state['short']}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        use_container_width=True,
                        on_click=reset_mobility_points,
                        args=(
                            removed_key,
                            force_key,
                            (
                                state["peak_key"], state["on_key"],
                                state["off_key"], state["vth_key"],
                                state["ss_key"], remove_key,
                            ),
                        ),
                    )
                    container.caption(
                        f"Removed: {len(st.session_state.get(removed_key, []))}"
                    )

            # ====================================================
            # ====================================================
            # ====================================================
            # ====================================================
            # Selected-direction controls directly below plots
            # ====================================================
            st.markdown(
                "<div class='compact-slider-area'></div>",
                unsafe_allow_html=True,
            )
            st.markdown(
                "<div class='slider-row-gap'></div>",
                unsafe_allow_html=True,
            )
            control_columns = st.columns([1, 1, 1, 1], gap="large")

            active_removed_key = (
                keys["removed_fwd"]
                if selected_direction == "Forward"
                else keys["removed_bwd"]
            )
            active_force_key = (
                keys["force_auto_peak_fwd"]
                if selected_direction == "Forward"
                else keys["force_auto_peak_bwd"]
            )

            def render_remove_buttons(
                parent,
                target_vg,
                prefix,
                reset_keys,
                state_key=None,
            ):
                if state_key:
                    _, target_row = snapped_slider_point(
                        active_state["df"], state_key
                    )
                else:
                    _, target_row = nearest_row_by_vg(
                        active_state["df"], target_vg
                    )
                remove_col, reset_col = parent.columns(2, gap="small")
                remove_col.button(
                    "✕",
                    key=f"{prefix}_remove",
                    use_container_width=True,
                    help="Remove selected point",
                    on_click=remove_mobility_point,
                    args=(
                        active_removed_key,
                        int(target_row["__source_index"]),
                        active_force_key,
                        reset_keys,
                    ),
                )
                reset_col.button(
                    "↶",
                    key=f"{prefix}_reset",
                    use_container_width=True,
                    help="Restore all removed points",
                    on_click=reset_mobility_points,
                    args=(active_removed_key, active_force_key, reset_keys),
                )

            # Column 1: Transfer Log controls.
            with control_columns[0]:
                st.markdown(
                    f"<div class='slider-heading' style='color:{direction_color};'>"
                    f"{selected_direction} · Transfer (Log)</div>",
                    unsafe_allow_html=True,
                )
                on_control_col, off_control_col = control_columns[0].columns(
                    2, gap="medium"
                )
                with on_control_col:
                    st.markdown(
                        "<div class='slider-heading'>ON</div>",
                        unsafe_allow_html=True,
                    )
                    render_discrete_vg_control(
                        "", "", active_state["on_key"], active_state["df"],
                        float(active_state["df"]["GateV"].iloc[active_state["on_auto_idx"]]),
                        f"active_on_{selected_direction}_{file_id}_{selected_sheet}_{operating_mode}",
                        on_control_col,
                    )
                    on_control_col.button(
                        "Auto Set",
                        key=f"active_on_auto_{selected_direction}_{file_id}_{selected_sheet}_{operating_mode}",
                        use_container_width=True,
                        on_click=set_state_value,
                        args=(active_state["on_key"], float(active_state["df"]["GateV"].iloc[active_state["on_auto_idx"]])),
                    )
                with off_control_col:
                    st.markdown(
                        "<div class='slider-heading'>OFF</div>",
                        unsafe_allow_html=True,
                    )
                    render_discrete_vg_control(
                        "", "", active_state["off_key"], active_state["df"],
                        float(active_state["df"]["GateV"].iloc[active_state["off_auto_idx"]]),
                        f"active_off_{selected_direction}_{file_id}_{selected_sheet}_{operating_mode}",
                        off_control_col,
                    )
                    off_control_col.button(
                        "Auto Set",
                        key=f"active_off_auto_{selected_direction}_{file_id}_{selected_sheet}_{operating_mode}",
                        use_container_width=True,
                        on_click=set_state_value,
                        args=(active_state["off_key"], float(active_state["df"]["GateV"].iloc[active_state["off_auto_idx"]])),
                    )

                st.markdown(
                    "<div class='control-section-spacer'></div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    "<div class='slider-heading'>Peak Elimination</div>",
                    unsafe_allow_html=True,
                )
                log_remove_vg = render_discrete_vg_control(
                    "", "", active_state["log_remove_key"], active_state["df"],
                    float(active_state["df"]["GateV"].iloc[active_state["auto_idx"]]),
                    f"log_peak_remove_{selected_direction}_{file_id}_{selected_sheet}_{operating_mode}",
                    control_columns[0],
                )
                render_remove_buttons(
                    control_columns[0],
                    log_remove_vg,
                    f"log_peak_{selected_direction}_{file_id}_{selected_sheet}_{operating_mode}",
                    (),
                    state_key=active_state["log_remove_key"],
                )

            # Column 2: SS Value slider with range-limited minimum detection.
            with control_columns[1]:
                st.markdown(
                    f"<div class='slider-heading' "
                    f"style='color:{direction_color};'>"
                    f"{selected_direction} · SS Value</div>",
                    unsafe_allow_html=True,
                )
                _, ss_inner, _ = control_columns[1].columns(
                    [0.08, 0.84, 0.08],
                    gap="small",
                )
                with ss_inner:
                    st.markdown(
                        "<div class='slider-heading'>SS</div>",
                        unsafe_allow_html=True,
                    )
                    render_discrete_vg_control(
                        "",
                        "",
                        active_state["ss_key"],
                        active_state["df"],
                        active_state["ss_default"],
                        (
                            f"active_ss_{selected_direction}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        ss_inner,
                    )
                    ss_inner.button(
                        "Auto Set",
                        key=(
                            f"active_ss_auto_{selected_direction}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        use_container_width=True,
                        on_click=set_state_value,
                        args=(
                            active_state["ss_key"],
                            float(active_state["ss_default"]),
                        ),
                    )
                    ss_range_cols = ss_inner.columns(2, gap="small")
                    render_persistent_number_input(
                        ss_range_cols[0],
                        "Vgs Start (V)",
                        active_state["ss_range_start_key"],
                        (
                            f"ss_start_widget_{selected_direction}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        active_state["ss_range_low"],
                        step=0.1,
                    )
                    render_persistent_number_input(
                        ss_range_cols[1],
                        "Vgs End (V)",
                        active_state["ss_range_end_key"],
                        (
                            f"ss_end_widget_{selected_direction}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        active_state["ss_range_high"],
                        step=0.1,
                    )

            # Column 3: Mobility + Peak Elimination.
            with control_columns[2]:
                st.markdown(
                    f"<div class='slider-heading' "
                    f"style='color:{direction_color};'>"
                    f"{selected_direction} · Mobility</div>",
                    unsafe_allow_html=True,
                )
                _, mobility_inner, _ = control_columns[2].columns(
                    [0.08, 0.84, 0.08],
                    gap="small",
                )
                with mobility_inner:
                    st.markdown(
                        "<div class='slider-heading'>Mobility</div>",
                        unsafe_allow_html=True,
                    )
                    render_discrete_vg_control(
                        "",
                        "",
                        active_state["peak_key"],
                        active_state["df"],
                        active_state["peak_default"],
                        (
                            f"active_mobility_{selected_direction}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        mobility_inner,
                    )
                    mobility_inner.button(
                        "Auto Set",
                        key=(
                            f"active_mobility_auto_{selected_direction}_"
                            f"{file_id}_{selected_sheet}_{operating_mode}"
                        ),
                        use_container_width=True,
                        on_click=set_state_value,
                        args=(
                            active_state["peak_key"],
                            float(active_state["peak_default"]),
                        ),
                    )

                    st.markdown(
                        "<div class='control-section-spacer'></div>",
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        "<div class='slider-heading'>Peak Elimination</div>",
                        unsafe_allow_html=True,
                    )
                    mobility_remove_vg = render_discrete_vg_control(
                        "",
                        "",
                        active_state["remove_key"],
                        active_state["df"],
                        float(
                            active_state["df"]["GateV"].iloc[
                                active_state["auto_idx"]
                            ]
                        ),
                        (
                            f"mobility_peak_remove_{selected_direction}_"
                            f"{file_id}_{selected_sheet}_{operating_mode}"
                        ),
                        mobility_inner,
                    )
                    render_remove_buttons(
                        mobility_inner,
                        mobility_remove_vg,
                        (
                            f"mobility_peak_{selected_direction}_{file_id}_"
                            f"{selected_sheet}_{operating_mode}"
                        ),
                        (),
                        state_key=active_state["remove_key"],
                    )

            # Column 4: Transfer Linear has no independent slider.
            with control_columns[3]:
                st.markdown(
                    "<div class='control-placeholder'></div>",
                    unsafe_allow_html=True,
                )
